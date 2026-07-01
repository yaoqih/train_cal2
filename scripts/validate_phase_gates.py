#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from dataclasses import asdict, dataclass, field
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import load_workbook


DEFAULT_REPRESENTATIVE_CASES = (
    "0117Z",
    "0310W",
    "0103W",
    "0213W",
    "0306W",
    "0128W",
    "0223W",
    "0130Z",
    "0201W",
)

HOOK_TOLERANCE_RATIO = 0.10
HOOK_TOLERANCE_MIN = 2
PHASES = ("H1", "H2", "H3", "H4", "H5")
PHASE_RANK = {phase: index for index, phase in enumerate(PHASES, start=1)}
VALID_TRANSITION_TYPES = {"enter", "exit", "skip", "fail", "stay"}
VALID_EDGE_STATUSES = {"SEED", "FORMING", "PORT_READY", "ACCEPTED", "DIGESTING", "DONE"}
VALID_TARGET_CONTRACTS = {
    "YARD_REBALANCE",
    "FUNCTION_LINE_SERVICE",
    "PRE_REPAIR_STAGING",
    "DISPATCH_SHED_QUEUE",
    "LOCO_AREA_STAGING",
    "CUN4_PORT_SHAPING",
    "STRICT_RELEASE",
    "MACHINE_ACCEPT",
    "REPAIR_INBOUND",
    "DEPOT_OUTBOUND",
    "DEPOT_SLOT",
    "DEPOT_DIGEST",
    "TAIL_CLOSEOUT",
    "SPECIAL_REPAIR_PROCESS",
}
PHASE_ALLOWED_TARGETS = {
    "H1": {
        "YARD_REBALANCE",
        "FUNCTION_LINE_SERVICE",
        "PRE_REPAIR_STAGING",
        "DISPATCH_SHED_QUEUE",
        "LOCO_AREA_STAGING",
        "CUN4_PORT_SHAPING",
    },
    "H2": {"CUN4_PORT_SHAPING", "YARD_REBALANCE", "STRICT_RELEASE"},
    "H3": {"STRICT_RELEASE", "MACHINE_ACCEPT", "REPAIR_INBOUND"},
    "H4": {"REPAIR_INBOUND", "DEPOT_OUTBOUND", "DEPOT_SLOT", "DEPOT_DIGEST", "SPECIAL_REPAIR_PROCESS"},
    "H5": {"TAIL_CLOSEOUT", "LOCO_AREA_STAGING", "FUNCTION_LINE_SERVICE", "YARD_REBALANCE"},
}
VALID_CANDIDATE_ACTION_FAMILIES = VALID_TARGET_CONTRACTS
VALID_RESOURCE_REQUESTS = {
    "yard_track_access",
    "function_line_access",
    "pre_repair_line_access",
    "dispatch_shed_access",
    "loco_area_access",
    "cun4_north_port_access",
    "cun4_release_gate_and_loco_end",
    "machine_accept_gate_and_receiver_capacity",
    "depot_inbound_route_and_slot",
    "depot_outbound_route",
    "depot_slot_capacity",
    "depot_detach_order_and_slot",
    "tail_route_and_loco_return",
    "special_process_resource",
}
VALID_RESOURCE_STATUSES = {"available", "constrained", "waiting", "blocked"}
VALID_GATE_DECISIONS = {"accept", "reject", "defer"}
PHASE_ALLOWED_CANDIDATE_FAMILIES = {
    "H1": {
        "YARD_REBALANCE",
        "FUNCTION_LINE_SERVICE",
        "PRE_REPAIR_STAGING",
        "DISPATCH_SHED_QUEUE",
        "LOCO_AREA_STAGING",
        "CUN4_PORT_SHAPING",
    },
    "H2": {"CUN4_PORT_SHAPING", "YARD_REBALANCE", "STRICT_RELEASE"},
    "H3": {"STRICT_RELEASE", "MACHINE_ACCEPT", "REPAIR_INBOUND"},
    "H4": {"REPAIR_INBOUND", "DEPOT_OUTBOUND", "DEPOT_SLOT", "DEPOT_DIGEST", "SPECIAL_REPAIR_PROCESS"},
    "H5": {
        "TAIL_CLOSEOUT",
        "LOCO_AREA_STAGING",
        "FUNCTION_LINE_SERVICE",
        "YARD_REBALANCE",
        "PRE_REPAIR_STAGING",
        "DISPATCH_SHED_QUEUE",
    },
}


@dataclass(frozen=True)
class Hook:
    case_id: str
    step: int
    line_raw: str
    line: str
    method: str
    count: int | None
    note: str


@dataclass
class ManualPhaseAudit:
    case_id: str
    source_path: str
    observed_hook_count: int
    hook_count_confidence: str
    hook_tolerance: int
    soft_hook_upper_bound: int
    h1_start: int | None = None
    h1_end: int | None = None
    h2_start: int | None = None
    h2_end: int | None = None
    h3_start: int | None = None
    h3_end: int | None = None
    h4_start: int | None = None
    h4_end: int | None = None
    h5_start: int | None = None
    h5_end: int | None = None
    phase_path: str = ""
    variant: str = ""
    confidence: str = "medium"
    failures: list[str] = field(default_factory=list)
    signals: dict[str, int | None] = field(default_factory=dict)


@dataclass
class TruthCaseAudit:
    case_id: str
    source_path: str
    vehicle_count: int
    movable_vehicle_count: int
    effective_contract_coverage: float
    residual_vehicle_ratio: float
    depot_target_vehicle_count: int
    depot_initial_vehicle_count: int
    pre_repair_target_vehicle_count: int
    dispatch_shed_target_vehicle_count: int
    function_line_target_vehicle_count: int
    storage_target_vehicle_count: int
    cun4_target_vehicle_count: int
    special_process_vehicle_count: int
    line_alias_unknown_count: int
    passed_p1_floor: bool
    failures: list[str] = field(default_factory=list)


@dataclass
class ProcessAudit:
    case_id: str
    process: str
    status: str
    checked: bool
    metric_1_name: str = ""
    metric_1_value: str = ""
    metric_2_name: str = ""
    metric_2_value: str = ""
    metric_3_name: str = ""
    metric_3_value: str = ""
    failures: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class PhaseGateContract:
    case_id: str
    source_path: str
    phase: str
    variant: str
    expected: bool
    skip_allowed: bool
    entry_step: int | None
    exit_step: int | None
    manual_phase_hook_count: int | None
    soft_phase_hook_upper_bound: int | None
    hard_exit_signal: str
    hook_count_confidence: str
    skip_reason: str = ""
    compressed_with: str = ""


@dataclass
class PhaseTraceAudit:
    case_id: str
    trace_row_index: int
    step_index: int | None
    from_phase: str
    to_phase: str
    transition_type: str
    passed: bool
    failures: list[str] = field(default_factory=list)


@dataclass
class ActionTraceAudit:
    case_id: str
    trace_row_index: int
    step_index: int | None
    phase: str
    edge_status: str
    target_contract: str
    status: str
    failures: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class ManualCandidateBaseline:
    case_id: str
    phase: str
    action_family: str
    variant: str
    baseline_level: str
    manual_hook_count: int
    manual_signal_step: int | None
    hook_count_confidence: str


@dataclass
class CandidateTraceAudit:
    case_id: str
    trace_row_index: int
    step_index: int | None
    phase: str
    candidate_id: str
    action_family: str
    resource_request: str
    candidate_level: str
    candidate_status: str
    status: str
    failures: list[str] = field(default_factory=list)


@dataclass
class ResourceTraceAudit:
    case_id: str
    trace_row_index: int
    step_index: int | None
    candidate_id: str
    resource_request: str
    resource_status: str
    status: str
    failures: list[str] = field(default_factory=list)


@dataclass
class DeltaTraceAudit:
    case_id: str
    trace_row_index: int
    step_index: int | None
    candidate_id: str
    gate_decision: str
    hard_violation_count: int | None
    status: str
    failures: list[str] = field(default_factory=list)


@dataclass
class OptimizationTraceAudit:
    case_id: str
    trace_row_index: int
    step_index: int | None
    candidate_id: str
    hook_cost: int | None
    rank: int | None
    selected: bool
    status: str
    failures: list[str] = field(default_factory=list)


@dataclass
class StateUpdateTraceAudit:
    case_id: str
    trace_row_index: int
    step_index: int | None
    candidate_id: str
    hook_increment: int | None
    rebuild_status: str
    next_phase: str
    status: str
    failures: list[str] = field(default_factory=list)


@dataclass
class Summary:
    manual_case_count: int
    truth_case_count: int
    matched_case_count: int
    standard_chain_count: int
    depot_digest_only_count: int
    port_release_without_standard_accept_count: int
    low_signal_or_short_chain_count: int
    median_observed_hooks: float | None
    p1_passed_case_count: int
    p1_failed_case_count: int
    skipped_manual_file_count: int
    skipped_truth_file_count: int
    duplicate_manual_case_count: int
    phase_contract_row_count: int
    phase_trace_record_count: int
    phase_trace_failed_record_count: int
    process_audit_row_count: int
    process_failed_row_count: int
    process_blocked_row_count: int
    action_trace_record_count: int
    action_trace_failed_record_count: int
    candidate_baseline_row_count: int
    candidate_trace_record_count: int
    candidate_trace_failed_record_count: int
    resource_trace_record_count: int
    resource_trace_failed_record_count: int
    delta_trace_record_count: int
    delta_trace_failed_record_count: int
    optimization_trace_record_count: int
    optimization_trace_failed_record_count: int
    state_update_trace_record_count: int
    state_update_trace_failed_record_count: int


def case_id_from_path(path: Path) -> str:
    match = re.search(r"(\d{4}[ZWzw])", path.name)
    if not match:
        raise ValueError(f"Cannot infer case id from {path}")
    return match.group(1).upper()


def try_case_id_from_path(path: Path) -> str | None:
    match = re.search(r"(\d{4}[ZWzw])", path.name)
    return match.group(1).upper() if match else None


def normalize_line(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    aliases = {
        "机走北1线": "机北1",
        "机走北2线": "机北2",
        "机走北": "机走北",
        "机北3": "机走北",
        "机走线南": "机南",
        "机南": "机南",
        "洗罐油漆北": "洗油北",
        "洗油北": "洗油北",
        "洗罐线北": "洗罐线北",
        "洗北": "洗罐线北",
        "洗罐站": "洗罐站",
        "洗南": "洗罐站",
        "调梁线北": "调梁线北",
        "调北": "调梁线北",
        "调梁棚": "调梁棚",
        "调棚": "调梁棚",
        "存4线": "存4线",
        "存4北": "存4线",
        "存4线南": "存4南",
        "存4南": "存4南",
        "存5北": "存5线北",
        "存5南": "存5线南",
        "存5线北": "存5线北",
        "存5线南": "存5线南",
        "联7线": "联7",
        "机库": "机库线",
        "库": "机库线",
        "注意库": "机库线",
        "修1": "修1库内",
        "修2": "修2库内",
        "修3": "修3库内",
        "修4": "修4库内",
        "存1": "存1线",
        "存2": "存2线",
        "注意存2": "存2线",
        "存2叉": "存2线",
        "存3": "存3线",
        "存4": "存4线",
        "注意存4": "存4线",
        "存5": "存5线",
        "注意存5": "存5线",
        "调": "调梁棚",
        "机": "机库线",
        "注意机": "机库线",
        "机走": "机走棚",
        "洗": "洗罐站",
        "油": "油漆线",
        "抛": "抛丸线",
        "轮": "卸轮线",
    }
    if text in aliases:
        return aliases[text]
    text = text.replace("線", "线")
    text = text.replace("库外", "库外")
    return text


def normalize_method(value: Any) -> str:
    return str(value or "").strip()


def normalize_note(value: Any) -> str:
    return str(value or "").strip()


def parse_int(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def is_effective_hook(line: Any, method: Any, count: int | None, note: Any) -> bool:
    return bool(
        str(line or "").strip()
        and (str(method or "").strip() or count is not None or str(note or "").strip())
    )


def parse_manual_hooks(path: Path) -> list[Hook]:
    case_id = case_id_from_path(path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    hooks: list[Hook] = []

    for row in sheet.iter_rows(values_only=True):
        for offset in (0, 5):
            if len(row) <= offset + 4:
                continue
            step = parse_int(row[offset])
            if step is None:
                continue
            line_raw = str(row[offset + 1] or "").strip()
            method = normalize_method(row[offset + 2])
            count_raw = row[offset + 3]
            count = parse_int(count_raw)
            note = normalize_note(row[offset + 4])
            if count is None and count_raw is not None and not note:
                note = normalize_note(count_raw)
            if not is_effective_hook(line_raw, method, count, note):
                continue
            hooks.append(
                Hook(
                    case_id=case_id,
                    step=step,
                    line_raw=line_raw,
                    line=normalize_line(line_raw),
                    method=method,
                    count=count,
                    note=note,
                )
            )

    return sorted(hooks, key=lambda item: item.step)


def first_step(hooks: list[Hook], predicate) -> int | None:
    for hook in hooks:
        if predicate(hook):
            return hook.step
    return None


def first_step_after(hooks: list[Hook], step: int | None, predicate) -> int | None:
    if step is None:
        return None
    return first_step([hook for hook in hooks if hook.step >= step], predicate)


def first_step_between(hooks: list[Hook], start: int | None, end: int | None, predicate) -> int | None:
    if start is None:
        return None
    return first_step(
        [hook for hook in hooks if hook.step >= start and (end is None or hook.step <= end)],
        predicate,
    )


def last_step(hooks: list[Hook], predicate) -> int | None:
    for hook in reversed(hooks):
        if predicate(hook):
            return hook.step
    return None


def last_step_after(hooks: list[Hook], step: int | None, predicate) -> int | None:
    if step is None:
        return None
    return last_step([hook for hook in hooks if hook.step >= step], predicate)


def is_front_service(hook: Hook) -> bool:
    return hook.line in {
        "存1线",
        "存2线",
        "存3线",
        "存5线",
        "调梁棚",
        "预修线",
        "洗罐站",
        "油漆线",
        "抛丸线",
        "卸轮线",
        "机库线",
        "机走棚",
    }


def is_cun4_shape(hook: Hook) -> bool:
    if hook.line != "存4线":
        return False
    text = f"{hook.method}{hook.note}"
    return any(token in text for token in ("北头", "代", "靠", "注意")) and not is_cun4_release(hook)


def is_cun4_release(hook: Hook) -> bool:
    return hook.line == "存4线" and hook.method == "-" and hook.count is not None and "北头摘" in hook.note


def is_machine_accept(hook: Hook) -> bool:
    return hook.line == "机库线" and hook.method == "+" and hook.count is not None and "接" in hook.note


def is_nonstandard_machine_accept(hook: Hook) -> bool:
    return hook.line == "机库线" and hook.method == "+" and hook.count is not None and "接" not in hook.note


def is_depot_digest(hook: Hook) -> bool:
    if hook.line not in {"修1库内", "修2库内", "修3库内", "修4库内"}:
        return False
    text = f"{hook.method}{hook.note}"
    return any(token in text for token in ("-", "摘", "库外"))


def is_depot_entry_or_digest_signal(hook: Hook) -> bool:
    return hook.line in {"修1库内", "修2库内", "修3库内", "修4库内"}


def is_tail_closeout(hook: Hook) -> bool:
    text = f"{hook.line_raw}{hook.method}{hook.note}"
    return "库回" in text or (hook.line in {"机库线", "大库"} and "回" in text)


def classify_variant(signals: dict[str, int | None], observed_hook_count: int) -> str:
    has_release = signals["cun4_release"] is not None
    has_accept = signals["machine_accept_after_release"] is not None
    has_digest = signals["depot_digest"] is not None
    has_depot_entry = signals["depot_entry"] is not None
    has_nonstandard_accept = signals["machine_nonstandard_after_release"] is not None
    if has_release and has_accept and has_digest:
        return "FULL_CHAIN_REPAIR"
    if observed_hook_count <= 10 and has_depot_entry and not has_accept and not has_nonstandard_accept:
        return "DIRECT_REPAIR_ENTRY"
    if has_digest and not has_release and not has_accept:
        return "DEPOT_DIGEST_ONLY"
    if has_release and not has_accept:
        return "MIXED_SIGNAL_REPAIR"
    return "GENERAL_ORGANIZATION"


def hook_tolerance(observed_hook_count: int) -> int:
    return max(HOOK_TOLERANCE_MIN, round(observed_hook_count * HOOK_TOLERANCE_RATIO))


def infer_manual_phase_audit(hooks: list[Hook]) -> ManualPhaseAudit:
    case_id = hooks[0].case_id if hooks else ""
    observed = len(hooks)
    tolerance = hook_tolerance(observed)
    signals = {
        "front_service": first_step(hooks, is_front_service),
        "cun4_shape": first_step(hooks, is_cun4_shape),
        "cun4_release": first_step(hooks, is_cun4_release),
        "machine_accept": first_step(hooks, is_machine_accept),
        "depot_entry": first_step(hooks, is_depot_entry_or_digest_signal),
        "depot_digest": first_step(hooks, is_depot_digest),
        "tail_closeout": first_step(hooks, is_tail_closeout),
    }
    signals["machine_accept_after_release"] = first_step_after(
        hooks,
        signals["cun4_release"],
        is_machine_accept,
    )
    signals["machine_nonstandard_after_release"] = first_step_after(
        hooks,
        signals["cun4_release"],
        is_nonstandard_machine_accept,
    )
    signals["depot_digest_after_accept"] = first_step_after(
        hooks,
        signals["machine_accept"],
        is_depot_digest,
    )
    signals["tail_closeout_after_digest"] = first_step_after(
        hooks,
        signals["depot_digest_after_accept"] or signals["depot_digest"],
        is_tail_closeout,
    )
    variant = classify_variant(signals, observed)

    audit = ManualPhaseAudit(
        case_id=case_id,
        source_path="",
        observed_hook_count=observed,
        hook_count_confidence="soft_manual_plan_may_omit_minor_hooks",
        hook_tolerance=tolerance,
        soft_hook_upper_bound=observed + tolerance,
        variant=variant,
        signals=signals,
    )

    if variant == "FULL_CHAIN_REPAIR":
        audit.h1_start = 1
        audit.h1_end = max(1, (signals["cun4_shape"] or signals["cun4_release"] or 1) - 1)
        audit.h2_start = signals["cun4_shape"] or audit.h1_end + 1
        audit.h2_end = max(audit.h2_start, (signals["cun4_release"] or audit.h2_start) - 1)
        audit.h3_start = signals["cun4_release"]
        audit.h3_end = signals["machine_accept_after_release"]
        audit.h4_start = signals["depot_digest_after_accept"] or signals["machine_accept_after_release"]
        audit.h4_end = last_step_after(hooks, signals["machine_accept_after_release"], is_depot_digest) or last_step(hooks, is_depot_digest)
        audit.h5_start = signals["tail_closeout_after_digest"] or ((audit.h4_end or observed) + 1)
        audit.h5_end = observed
        audit.phase_path = "H1->H2->H3->H4->H5"
    elif variant == "DEPOT_DIGEST_ONLY":
        audit.h4_start = signals["depot_digest"] or signals["depot_entry"] or 1
        audit.h4_end = last_step(hooks, is_depot_digest) or audit.h4_start
        if audit.h4_start and audit.h4_start > 1:
            audit.h1_start = 1
            audit.h1_end = audit.h4_start - 1
        audit.h5_start = signals["tail_closeout_after_digest"] or min(observed, audit.h4_end + 1)
        audit.h5_end = observed
        audit.phase_path = "H1->H4->H5" if audit.h1_start is not None else "H4->H5"
    elif variant == "DIRECT_REPAIR_ENTRY":
        audit.h3_start = signals["machine_accept_after_release"] or signals["machine_accept"] or signals["cun4_release"]
        audit.h3_end = signals["machine_accept_after_release"] or signals["machine_accept"] or signals["cun4_release"]
        audit.h4_start = signals["depot_digest"] or signals["depot_entry"] or audit.h3_end
        audit.h4_end = last_step(hooks, is_depot_digest) or audit.h4_start
        audit.h5_start = signals["tail_closeout_after_digest"] or signals["tail_closeout"] or min(observed, (audit.h4_end or observed) + 1)
        audit.h5_end = observed
        audit.phase_path = "H3/H4->H5"
    elif variant == "MIXED_SIGNAL_REPAIR":
        audit.h2_start = signals["cun4_shape"] or signals["cun4_release"]
        audit.h2_end = signals["cun4_release"]
        digest_start = signals["depot_digest"] or signals["depot_entry"]
        if digest_start is not None:
            audit.h4_start = digest_start
            audit.h4_end = last_step_after(hooks, digest_start, is_depot_digest) or digest_start
        audit.h5_start = signals["tail_closeout_after_digest"] or signals["tail_closeout"]
        if audit.h5_start is None and audit.h4_end is not None and audit.h4_end < observed:
            audit.h5_start = audit.h4_end + 1
        elif audit.h5_start is None and audit.h2_end is not None and audit.h2_end < observed:
            audit.h5_start = audit.h2_end + 1
        audit.h5_end = observed
        audit.phase_path = "H2->H4->H5" if audit.h4_start is not None else "H2->conservative->H5"
    else:
        audit.h1_start = 1 if observed else None
        audit.h1_end = observed
        audit.phase_path = "H1"

    validate_manual_phase_audit(audit)
    return audit


def validate_manual_phase_audit(audit: ManualPhaseAudit) -> None:
    failures: list[str] = []
    signals = audit.signals

    if audit.observed_hook_count == 0:
        failures.append("MANUAL_EMPTY")

    if audit.variant == "FULL_CHAIN_REPAIR":
        release = signals["cun4_release"]
        accept = signals["machine_accept_after_release"]
        digest = signals["depot_digest_after_accept"]
        if release is None:
            failures.append("H3_RELEASE_MISSING")
        if accept is None:
            failures.append("H3_ACCEPT_MISSING")
        if release is not None and accept is not None and accept < release:
            failures.append("H3_ACCEPT_BEFORE_RELEASE")
        if accept is not None and digest is None:
            failures.append("H4_DIGEST_AFTER_ACCEPT_MISSING")

    if audit.variant == "DEPOT_DIGEST_ONLY" and signals["depot_digest"] is None:
        failures.append("H4_DIGEST_MISSING")

    tail = signals["tail_closeout_after_digest"] or signals["tail_closeout"]
    digest_end = audit.h4_end
    if tail is not None and digest_end is not None and tail < digest_end:
        failures.append("H5_CLOSEOUT_BEFORE_DIGEST_DONE")

    if audit.variant == "MIXED_SIGNAL_REPAIR" and signals["machine_accept"] is not None:
        failures.append("MIXED_SIGNAL_HAS_STANDARD_ACCEPT")

    audit.failures = failures
    audit.confidence = "high" if not failures else "low"


def phase_bounds(audit: ManualPhaseAudit, phase: str) -> tuple[int | None, int | None]:
    phase_lower = phase.lower()
    return (
        getattr(audit, f"{phase_lower}_start"),
        getattr(audit, f"{phase_lower}_end"),
    )


def phase_exit_signal(audit: ManualPhaseAudit, phase: str) -> str:
    signals = audit.signals
    if phase == "H1":
        return "non_depot_progress_or_owner_ready"
    if phase == "H2":
        return "cun4_port_shape_ready"
    if phase == "H3":
        return "machine_accept_after_release" if signals.get("machine_accept_after_release") else "phase_skipped_or_compressed"
    if phase == "H4":
        return "depot_digest_after_accept_or_digest_only"
    if phase == "H5":
        return "tail_closeout_or_final_target_satisfied"
    return ""


def phase_skip_reason(audit: ManualPhaseAudit, phase: str) -> str:
    variant = audit.variant
    if variant == "FULL_CHAIN_REPAIR":
        return ""
    if variant == "DEPOT_DIGEST_ONLY" and phase in {"H1", "H2", "H3"}:
        return "depot_digest_only_starts_from_H4"
    if variant == "DIRECT_REPAIR_ENTRY" and phase in {"H1", "H2"}:
        return "direct_repair_entry_skips_front_and_port_shaping"
    if variant == "MIXED_SIGNAL_REPAIR" and phase == "H3":
        return "low_signal_or_nonstandard_accept_uses_conservative_path"
    if variant == "GENERAL_ORGANIZATION" and phase in {"H2", "H3", "H4", "H5"}:
        return "no_stable_phase_signal_in_manual_plan"
    return ""


def phase_compression(audit: ManualPhaseAudit, phase: str) -> str:
    if audit.variant == "DIRECT_REPAIR_ENTRY" and phase in {"H3", "H4"}:
        return "H3/H4"
    return ""


def build_phase_gate_contracts(manual_audits: list[ManualPhaseAudit]) -> list[PhaseGateContract]:
    contracts: list[PhaseGateContract] = []
    for audit in manual_audits:
        for phase in PHASES:
            start, end = phase_bounds(audit, phase)
            expected = start is not None or end is not None
            phase_hooks = None
            soft_upper = None
            if start is not None and end is not None:
                phase_hooks = max(0, end - start + 1)
                soft_upper = phase_hooks + hook_tolerance(phase_hooks)
            skip_reason = phase_skip_reason(audit, phase)
            contracts.append(
                PhaseGateContract(
                    case_id=audit.case_id,
                    source_path=audit.source_path,
                    phase=phase,
                    variant=audit.variant,
                    expected=expected,
                    skip_allowed=bool(skip_reason),
                    entry_step=start,
                    exit_step=end,
                    manual_phase_hook_count=phase_hooks,
                    soft_phase_hook_upper_bound=soft_upper,
                    hard_exit_signal=phase_exit_signal(audit, phase),
                    hook_count_confidence=audit.hook_count_confidence,
                    skip_reason=skip_reason,
                    compressed_with=phase_compression(audit, phase),
                )
            )
    return contracts


def phase_for_manual_step(audit: ManualPhaseAudit, step: int) -> str:
    for phase in PHASES:
        start, end = phase_bounds(audit, phase)
        if start is not None and end is not None and start <= step <= end:
            return phase
    for phase in PHASES:
        start, end = phase_bounds(audit, phase)
        if start is not None and step >= start and (end is None or step <= end):
            return phase
    return ""


def manual_hook_action_family(hook: Hook) -> str:
    if is_cun4_release(hook):
        return "STRICT_RELEASE"
    if is_machine_accept(hook) or is_nonstandard_machine_accept(hook):
        return "MACHINE_ACCEPT"
    if is_cun4_shape(hook):
        return "CUN4_PORT_SHAPING"
    if is_depot_digest(hook):
        return "DEPOT_DIGEST"
    if is_tail_closeout(hook):
        return "TAIL_CLOSEOUT"
    if hook.line in {"修1库内", "修2库内", "修3库内", "修4库内"}:
        return "REPAIR_INBOUND"
    if hook.line == "预修线":
        return "PRE_REPAIR_STAGING"
    if hook.line == "调梁棚":
        return "DISPATCH_SHED_QUEUE"
    if hook.line in {"洗罐站", "油漆线", "抛丸线", "卸轮线", "机库线", "机走棚"}:
        return "FUNCTION_LINE_SERVICE"
    if hook.line in {"存1线", "存2线", "存3线", "存4线", "存5线"}:
        return "YARD_REBALANCE"
    if hook.line in {"机库线", "大库"} and is_tail_closeout(hook):
        return "TAIL_CLOSEOUT"
    return "UNKNOWN"


def candidate_baseline_level(audit: ManualPhaseAudit, phase: str, action_family: str) -> str:
    if audit.variant == "FULL_CHAIN_REPAIR":
        if phase == "H3" and action_family in {"STRICT_RELEASE", "MACHINE_ACCEPT"}:
            return "critical"
        if phase == "H4" and action_family == "DEPOT_DIGEST":
            return "critical"
        if phase == "H2" and action_family == "CUN4_PORT_SHAPING":
            return "important"
    if audit.variant == "DEPOT_DIGEST_ONLY" and phase == "H4" and action_family == "DEPOT_DIGEST":
        return "critical"
    if audit.variant == "MIXED_SIGNAL_REPAIR" and phase == "H2" and action_family in {"STRICT_RELEASE", "CUN4_PORT_SHAPING"}:
        return "important"
    if audit.variant == "DIRECT_REPAIR_ENTRY" and phase in {"H3", "H4"} and action_family in {"STRICT_RELEASE", "REPAIR_INBOUND", "DEPOT_DIGEST"}:
        return "critical"
    return "observed"


def build_manual_candidate_baselines(
    manual_audits: list[ManualPhaseAudit],
    hooks_by_case: dict[str, list[Hook]],
) -> list[ManualCandidateBaseline]:
    baselines: list[ManualCandidateBaseline] = []
    seen: set[tuple[str, str, str]] = set()
    for audit in manual_audits:
        hooks = hooks_by_case.get(audit.case_id, [])
        manual_counts: Counter[tuple[str, str]] = Counter()
        first_steps: dict[tuple[str, str], int] = {}
        for hook in hooks:
            phase = phase_for_manual_step(audit, hook.step)
            action_family = manual_hook_action_family(hook)
            if not phase or action_family == "UNKNOWN":
                continue
            key = (phase, action_family)
            manual_counts[key] += 1
            first_steps.setdefault(key, hook.step)

        required_pairs: set[tuple[str, str]] = set()
        if audit.variant == "FULL_CHAIN_REPAIR":
            required_pairs.update(
                {
                    ("H2", "CUN4_PORT_SHAPING"),
                    ("H3", "STRICT_RELEASE"),
                    ("H3", "MACHINE_ACCEPT"),
                    ("H4", "DEPOT_DIGEST"),
                }
            )
        elif audit.variant == "DEPOT_DIGEST_ONLY":
            required_pairs.add(("H4", "DEPOT_DIGEST"))
        elif audit.variant == "MIXED_SIGNAL_REPAIR":
            required_pairs.add(("H2", "STRICT_RELEASE"))
            if audit.signals.get("cun4_shape") is not None:
                required_pairs.add(("H2", "CUN4_PORT_SHAPING"))
        elif audit.variant == "DIRECT_REPAIR_ENTRY":
            required_pairs.add(("H3", "STRICT_RELEASE"))
            required_pairs.add(("H4", "REPAIR_INBOUND"))

        for phase, action_family in sorted(set(manual_counts) | required_pairs):
            key = (audit.case_id, phase, action_family)
            if key in seen:
                continue
            seen.add(key)
            baselines.append(
                ManualCandidateBaseline(
                    case_id=audit.case_id,
                    phase=phase,
                    action_family=action_family,
                    variant=audit.variant,
                    baseline_level=candidate_baseline_level(audit, phase, action_family),
                    manual_hook_count=manual_counts[(phase, action_family)],
                    manual_signal_step=first_steps.get((phase, action_family)),
                    hook_count_confidence=audit.hook_count_confidence,
                )
            )
    return baselines


def read_trace_records(path: Path, keys: tuple[str, ...]) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(path)
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8", newline="") as file:
            return list(csv.DictReader(file))
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in keys:
            value = payload.get(key)
            if isinstance(value, list):
                return value
    raise ValueError(f"Unsupported trace format: {path}")


def read_phase_trace_records(path: Path) -> list[dict[str, Any]]:
    return read_trace_records(
        path,
        ("phase_gate_records", "PhaseGateRecord", "phaseTrace", "records", "rows"),
    )


def read_action_trace_records(path: Path) -> list[dict[str, Any]]:
    return read_trace_records(
        path,
        ("action_trace_records", "ActionTraceRecord", "actionTrace", "records", "rows"),
    )


def read_candidate_trace_records(path: Path) -> list[dict[str, Any]]:
    return read_trace_records(
        path,
        ("candidate_trace_records", "CandidateTraceRecord", "candidateTrace", "records", "rows"),
    )


def read_resource_trace_records(path: Path) -> list[dict[str, Any]]:
    return read_trace_records(
        path,
        ("resource_trace_records", "ResourceTraceRecord", "resourceTrace", "records", "rows"),
    )


def read_delta_trace_records(path: Path) -> list[dict[str, Any]]:
    return read_trace_records(
        path,
        ("delta_trace_records", "DeltaTraceRecord", "deltaTrace", "records", "rows"),
    )


def read_optimization_trace_records(path: Path) -> list[dict[str, Any]]:
    return read_trace_records(
        path,
        ("optimization_trace_records", "OptimizationTraceRecord", "optimizationTrace", "records", "rows"),
    )


def read_state_update_trace_records(path: Path) -> list[dict[str, Any]]:
    return read_trace_records(
        path,
        ("state_update_trace_records", "StateUpdateTraceRecord", "stateUpdateTrace", "records", "rows"),
    )


def normalize_phase(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text in PHASE_RANK:
        return text
    match = re.search(r"H[1-5]", text)
    return match.group(0) if match else text


def parse_optional_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    parsed = parse_int(value)
    if parsed is not None:
        return parsed
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def first_present(record: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in record and record[key] not in (None, ""):
            return record[key]
    return ""


def audit_phase_trace_records(
    records: list[dict[str, Any]],
    contracts: list[PhaseGateContract],
) -> list[PhaseTraceAudit]:
    audits: list[PhaseTraceAudit] = []
    active_phase_by_case: dict[str, str] = {}
    contract_by_case_phase = {
        (contract.case_id, contract.phase): contract for contract in contracts
    }
    for index, record in enumerate(records, start=1):
        case_id = str(record.get("case_id") or record.get("caseId") or "").upper()
        from_phase = normalize_phase(record.get("from_phase") or record.get("fromPhase"))
        to_phase = normalize_phase(record.get("to_phase") or record.get("toPhase"))
        transition_type = str(
            record.get("transition_type") or record.get("transitionType") or ""
        ).strip().lower()
        step_index = parse_optional_int(record.get("step_index") or record.get("stepIndex"))
        failures: list[str] = []

        if not case_id:
            failures.append("TRACE_CASE_ID_MISSING")
        if transition_type not in VALID_TRANSITION_TYPES:
            failures.append("TRACE_TRANSITION_TYPE_INVALID")
        if transition_type != "enter" and not from_phase:
            failures.append("TRACE_FROM_PHASE_MISSING")
        if transition_type != "exit" and not to_phase:
            failures.append("TRACE_TO_PHASE_MISSING")
        if from_phase and from_phase not in PHASE_RANK:
            failures.append("TRACE_FROM_PHASE_UNKNOWN")
        if to_phase and to_phase not in PHASE_RANK:
            failures.append("TRACE_TO_PHASE_UNKNOWN")

        if not failures and case_id:
            if to_phase:
                contract = contract_by_case_phase.get((case_id, to_phase))
                if contract is None:
                    failures.append("TRACE_PHASE_CONTRACT_MISSING")
                elif transition_type == "skip" and not contract.skip_allowed and not contract.compressed_with:
                    failures.append("TRACE_SKIP_NOT_ALLOWED_BY_CONTRACT")
                elif transition_type in {"enter", "stay", "exit"} and not contract.expected and not contract.compressed_with:
                    failures.append("TRACE_PHASE_NOT_EXPECTED_BY_CONTRACT")
            active = active_phase_by_case.get(case_id)
            if transition_type == "enter":
                if active and active != to_phase:
                    failures.append("TRACE_ENTER_WITH_ACTIVE_PHASE")
                active_phase_by_case[case_id] = to_phase
            elif transition_type == "stay":
                if active and from_phase and active != from_phase:
                    failures.append("TRACE_STAY_FROM_PHASE_NOT_ACTIVE")
                if from_phase and to_phase and from_phase != to_phase:
                    failures.append("TRACE_STAY_PHASE_CHANGED")
                active_phase_by_case.setdefault(case_id, from_phase or to_phase)
            elif transition_type in {"exit", "skip", "fail"}:
                if active and from_phase and active != from_phase:
                    failures.append("TRACE_EXIT_FROM_PHASE_NOT_ACTIVE")
                if transition_type == "skip":
                    from_rank = PHASE_RANK.get(from_phase)
                    to_rank = PHASE_RANK.get(to_phase)
                    if from_rank and to_rank and to_rank <= from_rank:
                        failures.append("TRACE_SKIP_NOT_FORWARD")
                active_phase_by_case[case_id] = to_phase or active_phase_by_case.get(case_id, "")

        audits.append(
            PhaseTraceAudit(
                case_id=case_id,
                trace_row_index=index,
                step_index=step_index,
                from_phase=from_phase,
                to_phase=to_phase,
                transition_type=transition_type,
                passed=not failures,
                failures=failures,
            )
        )
    return audits


def audit_action_trace_records(
    records: list[dict[str, Any]],
    contracts: list[PhaseGateContract],
) -> list[ActionTraceAudit]:
    audits: list[ActionTraceAudit] = []
    contract_by_case_phase = {
        (contract.case_id, contract.phase): contract for contract in contracts
    }
    for index, record in enumerate(records, start=1):
        case_id = str(first_present(record, "case_id", "caseId") or "").upper()
        step_index = parse_optional_int(first_present(record, "step_index", "stepIndex"))
        phase = normalize_phase(first_present(record, "phase", "current_phase", "currentPhase"))
        edge_id = str(first_present(record, "edge_id", "edgeId") or "").strip()
        edge_status = str(first_present(record, "edge_status", "edgeStatus") or "").strip().upper()
        contract_id = str(first_present(record, "contract_id", "contractId") or "").strip()
        contract_variant = str(first_present(record, "contract_variant", "variant", "contractVariant") or "").strip()
        target_contract = str(first_present(record, "target_contract", "targetContract") or "").strip().upper()
        target_reason = str(first_present(record, "target_reason", "targetReason") or "").strip()
        failures: list[str] = []

        if not case_id:
            failures.append("ACTION_CASE_ID_MISSING")
        if step_index is None:
            failures.append("ACTION_STEP_INDEX_MISSING")
        if not phase:
            failures.append("P2_PHASE_MISSING")
        elif phase not in PHASE_RANK:
            failures.append("P2_PHASE_UNKNOWN")
        if not edge_id:
            failures.append("P2_EDGE_ID_MISSING")
        if not edge_status:
            failures.append("P2_EDGE_STATUS_MISSING")
        elif edge_status not in VALID_EDGE_STATUSES:
            failures.append("P2_EDGE_STATUS_UNKNOWN")
        if not contract_id:
            failures.append("P3_CONTRACT_ID_MISSING")
        if not contract_variant:
            failures.append("P3_CONTRACT_VARIANT_MISSING")
        if not target_contract:
            failures.append("P4_TARGET_CONTRACT_MISSING")
        elif target_contract not in VALID_TARGET_CONTRACTS:
            failures.append("P4_TARGET_CONTRACT_UNKNOWN")
        if not target_reason:
            failures.append("P4_TARGET_REASON_MISSING")

        if case_id and phase and phase in PHASE_RANK:
            contract = contract_by_case_phase.get((case_id, phase))
            if contract is None:
                failures.append("ACTION_PHASE_CONTRACT_MISSING")
            elif not contract.expected and not contract.compressed_with:
                failures.append("ACTION_PHASE_NOT_EXPECTED_BY_CONTRACT")
            allowed_targets = PHASE_ALLOWED_TARGETS.get(phase, set())
            if target_contract and target_contract in VALID_TARGET_CONTRACTS and target_contract not in allowed_targets:
                failures.append("P4_TARGET_NOT_ALLOWED_IN_PHASE")

        audits.append(
            ActionTraceAudit(
                case_id=case_id,
                trace_row_index=index,
                step_index=step_index,
                phase=phase,
                edge_status=edge_status,
                target_contract=target_contract,
                status="passed" if not failures else "failed",
                failures=failures,
            )
        )
    return audits


def audit_candidate_trace_records(
    records: list[dict[str, Any]],
    contracts: list[PhaseGateContract],
    baselines: list[ManualCandidateBaseline],
) -> list[CandidateTraceAudit]:
    audits: list[CandidateTraceAudit] = []
    contract_by_case_phase = {
        (contract.case_id, contract.phase): contract for contract in contracts
    }
    baseline_keys = {
        (baseline.case_id, baseline.phase, baseline.action_family) for baseline in baselines
    }
    for index, record in enumerate(records, start=1):
        case_id = str(first_present(record, "case_id", "caseId") or "").upper()
        step_index = parse_optional_int(first_present(record, "step_index", "stepIndex"))
        phase = normalize_phase(first_present(record, "phase", "current_phase", "currentPhase"))
        candidate_id = str(first_present(record, "candidate_id", "candidateId") or "").strip()
        source_contract = str(first_present(record, "source_contract", "sourceContract", "contract_id", "contractId") or "").strip()
        action_family = str(first_present(record, "action_family", "actionFamily") or "").strip().upper()
        resource_request = str(first_present(record, "resource_request", "resourceRequest") or "").strip()
        why_generated = str(first_present(record, "why_generated", "whyGenerated") or "").strip()
        candidate_level = str(first_present(record, "candidate_level", "candidateLevel", "baseline_level", "baselineLevel") or "").strip().lower()
        candidate_status = str(first_present(record, "candidate_status", "candidateStatus") or "generated").strip().lower()
        failures: list[str] = []

        if not case_id:
            failures.append("CANDIDATE_CASE_ID_MISSING")
        if step_index is None:
            failures.append("CANDIDATE_STEP_INDEX_MISSING")
        if not phase:
            failures.append("CANDIDATE_PHASE_MISSING")
        elif phase not in PHASE_RANK:
            failures.append("CANDIDATE_PHASE_UNKNOWN")
        if not candidate_id:
            failures.append("CANDIDATE_ID_MISSING")
        if not source_contract:
            failures.append("CANDIDATE_SOURCE_CONTRACT_MISSING")
        if not action_family:
            failures.append("CANDIDATE_ACTION_FAMILY_MISSING")
        elif action_family not in VALID_CANDIDATE_ACTION_FAMILIES:
            failures.append("CANDIDATE_ACTION_FAMILY_UNKNOWN")
        if not resource_request:
            failures.append("CANDIDATE_RESOURCE_REQUEST_MISSING")
        elif resource_request not in VALID_RESOURCE_REQUESTS:
            failures.append("CANDIDATE_RESOURCE_REQUEST_UNKNOWN")
        if not why_generated:
            failures.append("CANDIDATE_WHY_GENERATED_MISSING")
        if candidate_level and candidate_level not in {"critical", "important", "structural", "observed"}:
            failures.append("CANDIDATE_LEVEL_UNKNOWN")
        if candidate_status not in {"generated", "kept", "rejected", "accepted"}:
            failures.append("CANDIDATE_STATUS_UNKNOWN")

        if case_id and phase in PHASE_RANK:
            contract = contract_by_case_phase.get((case_id, phase))
            if contract is None:
                failures.append("CANDIDATE_PHASE_CONTRACT_MISSING")
            elif not contract.expected and not contract.compressed_with:
                failures.append("CANDIDATE_PHASE_NOT_EXPECTED_BY_CONTRACT")
            allowed = PHASE_ALLOWED_CANDIDATE_FAMILIES.get(phase, set())
            if action_family in VALID_CANDIDATE_ACTION_FAMILIES and action_family not in allowed:
                if (case_id, phase, action_family) not in baseline_keys:
                    failures.append("CANDIDATE_ACTION_NOT_ALLOWED_IN_PHASE")

        audits.append(
            CandidateTraceAudit(
                case_id=case_id,
                trace_row_index=index,
                step_index=step_index,
                phase=phase,
                candidate_id=candidate_id,
                action_family=action_family,
                resource_request=resource_request,
                candidate_level=candidate_level,
                candidate_status=candidate_status,
                status="passed" if not failures else "failed",
                failures=failures,
            )
        )
    return audits


def audit_resource_trace_records(
    records: list[dict[str, Any]],
    candidate_audits: list[CandidateTraceAudit],
) -> list[ResourceTraceAudit]:
    audits: list[ResourceTraceAudit] = []
    known_candidates = {
        audit.candidate_id: audit
        for audit in candidate_audits
        if audit.candidate_id
    }
    for index, record in enumerate(records, start=1):
        case_id = str(first_present(record, "case_id", "caseId") or "").upper()
        step_index = parse_optional_int(first_present(record, "step_index", "stepIndex"))
        candidate_id = str(first_present(record, "candidate_id", "candidateId") or "").strip()
        resource_request = str(first_present(record, "resource_request", "resourceRequest") or "").strip()
        resource_scope = str(first_present(record, "resource_scope", "resourceScope") or "").strip()
        resource_status = str(first_present(record, "resource_status", "resourceStatus") or "").strip().lower()
        arbitration_reason = str(first_present(record, "arbitration_reason", "arbitrationReason") or "").strip()
        evidence_ids = str(first_present(record, "evidence_ids", "evidenceIds") or "").strip()
        blocker_ids = str(first_present(record, "blocker_ids", "blockerIds") or "").strip()
        failures: list[str] = []

        if not case_id:
            failures.append("RESOURCE_CASE_ID_MISSING")
        if step_index is None:
            failures.append("RESOURCE_STEP_INDEX_MISSING")
        if not candidate_id:
            failures.append("RESOURCE_CANDIDATE_ID_MISSING")
        elif candidate_id not in known_candidates:
            failures.append("RESOURCE_CANDIDATE_UNKNOWN")
        if not resource_request:
            failures.append("RESOURCE_REQUEST_MISSING")
        elif resource_request not in VALID_RESOURCE_REQUESTS:
            failures.append("RESOURCE_REQUEST_UNKNOWN")
        if candidate_id in known_candidates and resource_request:
            expected_request = known_candidates[candidate_id].resource_request
            if expected_request and resource_request != expected_request:
                failures.append("RESOURCE_REQUEST_MISMATCH")
        if not resource_scope:
            failures.append("RESOURCE_SCOPE_MISSING")
        if resource_status not in VALID_RESOURCE_STATUSES:
            failures.append("RESOURCE_STATUS_UNKNOWN")
        if not arbitration_reason:
            failures.append("RESOURCE_ARBITRATION_REASON_MISSING")
        if not evidence_ids:
            failures.append("RESOURCE_EVIDENCE_MISSING")
        if resource_status in {"blocked", "waiting"} and not blocker_ids:
            failures.append("RESOURCE_BLOCKER_MISSING")

        audits.append(
            ResourceTraceAudit(
                case_id=case_id,
                trace_row_index=index,
                step_index=step_index,
                candidate_id=candidate_id,
                resource_request=resource_request,
                resource_status=resource_status,
                status="passed" if not failures else "failed",
                failures=failures,
            )
        )
    return audits


def audit_delta_trace_records(
    records: list[dict[str, Any]],
    candidate_audits: list[CandidateTraceAudit],
    resource_audits: list[ResourceTraceAudit],
) -> list[DeltaTraceAudit]:
    audits: list[DeltaTraceAudit] = []
    known_candidates = {
        audit.candidate_id: audit
        for audit in candidate_audits
        if audit.candidate_id
    }
    resource_status_by_candidate = {
        audit.candidate_id: audit.resource_status
        for audit in resource_audits
        if audit.status == "passed"
    }
    for index, record in enumerate(records, start=1):
        case_id = str(first_present(record, "case_id", "caseId") or "").upper()
        step_index = parse_optional_int(first_present(record, "step_index", "stepIndex"))
        candidate_id = str(first_present(record, "candidate_id", "candidateId") or "").strip()
        contract_delta = str(first_present(record, "contract_delta", "contractDelta") or "").strip()
        resource_delta = str(first_present(record, "resource_delta", "resourceDelta") or "").strip()
        hard_violation_count = parse_optional_int(first_present(record, "hard_violation_count", "hardViolationCount"))
        hard_gate_reason = str(first_present(record, "hard_gate_reason", "hardGateReason") or "").strip()
        gate_decision = str(first_present(record, "gate_decision", "gateDecision") or "").strip().lower()
        evidence_ids = str(first_present(record, "evidence_ids", "evidenceIds") or "").strip()
        failures: list[str] = []

        if not case_id:
            failures.append("DELTA_CASE_ID_MISSING")
        if step_index is None:
            failures.append("DELTA_STEP_INDEX_MISSING")
        if not candidate_id:
            failures.append("DELTA_CANDIDATE_ID_MISSING")
        elif candidate_id not in known_candidates:
            failures.append("DELTA_CANDIDATE_UNKNOWN")
        if not contract_delta:
            failures.append("CONTRACT_DELTA_MISSING")
        if not resource_delta:
            failures.append("RESOURCE_DELTA_MISSING")
        if hard_violation_count is None:
            failures.append("HARD_VIOLATION_COUNT_MISSING")
        elif hard_violation_count < 0:
            failures.append("HARD_VIOLATION_COUNT_INVALID")
        if not hard_gate_reason:
            failures.append("HARD_GATE_REASON_MISSING")
        if gate_decision not in VALID_GATE_DECISIONS:
            failures.append("GATE_DECISION_UNKNOWN")
        if not evidence_ids:
            failures.append("DELTA_EVIDENCE_MISSING")

        if gate_decision == "accept" and hard_violation_count and hard_violation_count > 0:
            failures.append("HARD_VIOLATION_ACCEPTED")
        if gate_decision == "reject" and hard_violation_count == 0:
            failures.append("REJECT_WITHOUT_HARD_VIOLATION")
        resource_status = resource_status_by_candidate.get(candidate_id)
        if gate_decision == "accept" and resource_status in {"blocked", "waiting"}:
            failures.append("UNAVAILABLE_RESOURCE_ACCEPTED")
        if candidate_id in known_candidates:
            candidate = known_candidates[candidate_id]
            if candidate.candidate_level == "critical" and gate_decision == "reject" and hard_violation_count == 0:
                failures.append("CRITICAL_CANDIDATE_REJECTED_WITHOUT_HARD_REASON")

        audits.append(
            DeltaTraceAudit(
                case_id=case_id,
                trace_row_index=index,
                step_index=step_index,
                candidate_id=candidate_id,
                gate_decision=gate_decision,
                hard_violation_count=hard_violation_count,
                status="passed" if not failures else "failed",
                failures=failures,
            )
        )
    return audits


def parse_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    return None


def audit_optimization_trace_records(
    records: list[dict[str, Any]],
    candidate_audits: list[CandidateTraceAudit],
    delta_audits: list[DeltaTraceAudit],
) -> list[OptimizationTraceAudit]:
    audits: list[OptimizationTraceAudit] = []
    known_candidates = {
        audit.candidate_id: audit
        for audit in candidate_audits
        if audit.candidate_id
    }
    accepted_delta_candidates = {
        audit.candidate_id
        for audit in delta_audits
        if audit.status == "passed" and audit.gate_decision == "accept"
    }
    for index, record in enumerate(records, start=1):
        case_id = str(first_present(record, "case_id", "caseId") or "").upper()
        step_index = parse_optional_int(first_present(record, "step_index", "stepIndex"))
        candidate_id = str(first_present(record, "candidate_id", "candidateId") or "").strip()
        hook_cost = parse_optional_int(first_present(record, "hook_cost", "hookCost"))
        rank = parse_optional_int(first_present(record, "rank", "candidate_rank", "candidateRank"))
        selected_raw = first_present(record, "selected", "is_selected", "isSelected")
        selected = parse_bool(selected_raw)
        why_ranked = str(first_present(record, "why_ranked", "whyRanked") or "").strip()
        evidence_ids = str(first_present(record, "evidence_ids", "evidenceIds") or "").strip()
        failures: list[str] = []

        if not case_id:
            failures.append("OPT_CASE_ID_MISSING")
        if step_index is None:
            failures.append("OPT_STEP_INDEX_MISSING")
        if not candidate_id:
            failures.append("OPT_CANDIDATE_ID_MISSING")
        elif candidate_id not in known_candidates:
            failures.append("OPT_CANDIDATE_UNKNOWN")
        elif candidate_id not in accepted_delta_candidates:
            failures.append("OPT_CANDIDATE_NOT_ACCEPTED_BY_P7")
        if hook_cost is None:
            failures.append("OPT_HOOK_COST_MISSING")
        elif hook_cost <= 0:
            failures.append("OPT_HOOK_COST_INVALID")
        if rank is None:
            failures.append("OPT_RANK_MISSING")
        elif rank <= 0:
            failures.append("OPT_RANK_INVALID")
        if selected is None:
            failures.append("OPT_SELECTED_MISSING")
            selected_bool = False
        else:
            selected_bool = selected
        if not why_ranked:
            failures.append("OPT_WHY_RANKED_MISSING")
        if not evidence_ids:
            failures.append("OPT_EVIDENCE_MISSING")

        audits.append(
            OptimizationTraceAudit(
                case_id=case_id,
                trace_row_index=index,
                step_index=step_index,
                candidate_id=candidate_id,
                hook_cost=hook_cost,
                rank=rank,
                selected=selected_bool,
                status="passed" if not failures else "failed",
                failures=failures,
            )
        )
    return audits


def audit_state_update_trace_records(
    records: list[dict[str, Any]],
    optimization_audits: list[OptimizationTraceAudit],
) -> list[StateUpdateTraceAudit]:
    audits: list[StateUpdateTraceAudit] = []
    selected_candidates = {
        audit.candidate_id: audit
        for audit in optimization_audits
        if audit.status == "passed" and audit.selected
    }
    for index, record in enumerate(records, start=1):
        case_id = str(first_present(record, "case_id", "caseId") or "").upper()
        step_index = parse_optional_int(first_present(record, "step_index", "stepIndex"))
        candidate_id = str(first_present(record, "candidate_id", "candidateId") or "").strip()
        pre_state_signature = str(first_present(record, "pre_state_signature", "preStateSignature") or "").strip()
        post_state_signature = str(first_present(record, "post_state_signature", "postStateSignature") or "").strip()
        hook_increment = parse_optional_int(first_present(record, "hook_increment", "hookIncrement"))
        remaining_before = parse_optional_int(first_present(record, "remaining_obligation_before", "remainingObligationBefore"))
        remaining_after = parse_optional_int(first_present(record, "remaining_obligation_after", "remainingObligationAfter"))
        rebuild_status = str(first_present(record, "rebuild_status", "rebuildStatus") or "").strip().lower()
        next_phase = normalize_phase(first_present(record, "next_phase", "nextPhase"))
        evidence_ids = str(first_present(record, "evidence_ids", "evidenceIds") or "").strip()
        failures: list[str] = []

        if not case_id:
            failures.append("STATE_CASE_ID_MISSING")
        if step_index is None:
            failures.append("STATE_STEP_INDEX_MISSING")
        if not candidate_id:
            failures.append("STATE_CANDIDATE_ID_MISSING")
        elif candidate_id not in selected_candidates:
            failures.append("STATE_CANDIDATE_NOT_SELECTED_BY_P8")
        if not pre_state_signature:
            failures.append("PRE_STATE_SIGNATURE_MISSING")
        if not post_state_signature:
            failures.append("POST_STATE_SIGNATURE_MISSING")
        if pre_state_signature and post_state_signature and pre_state_signature == post_state_signature:
            failures.append("STATE_SIGNATURE_UNCHANGED")
        if hook_increment is None:
            failures.append("HOOK_INCREMENT_MISSING")
        elif hook_increment <= 0:
            failures.append("HOOK_INCREMENT_INVALID")
        if remaining_before is None:
            failures.append("REMAINING_OBLIGATION_BEFORE_MISSING")
        if remaining_after is None:
            failures.append("REMAINING_OBLIGATION_AFTER_MISSING")
        if remaining_before is not None and remaining_after is not None and remaining_after > remaining_before:
            failures.append("REMAINING_OBLIGATION_INCREASED")
        if rebuild_status != "success":
            failures.append("REBUILD_STATUS_NOT_SUCCESS")
        if not next_phase:
            failures.append("NEXT_PHASE_MISSING")
        elif next_phase not in PHASE_RANK:
            failures.append("NEXT_PHASE_UNKNOWN")
        if not evidence_ids:
            failures.append("STATE_EVIDENCE_MISSING")

        audits.append(
            StateUpdateTraceAudit(
                case_id=case_id,
                trace_row_index=index,
                step_index=step_index,
                candidate_id=candidate_id,
                hook_increment=hook_increment,
                rebuild_status=rebuild_status,
                next_phase=next_phase,
                status="passed" if not failures else "failed",
                failures=failures,
            )
        )
    return audits


def load_truth_case(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def target_lines(car: dict[str, Any]) -> list[str]:
    return [normalize_line(item) for item in car.get("TargetLines") or []]


def current_line(car: dict[str, Any]) -> str:
    return normalize_line(car.get("Line"))


def is_satisfied(car: dict[str, Any]) -> bool:
    line = current_line(car)
    return line in set(target_lines(car))


def contract_family(car: dict[str, Any]) -> str | None:
    targets = set(target_lines(car))
    line = current_line(car)
    repair_process = str(car.get("RepairProcess") or "")
    if targets & {"修1库内", "修2库内", "修3库内", "修4库内", "修1库外", "修2库外", "修3库外", "修4库外"}:
        return "REPAIR_DEPOT"
    if "预修线" in targets:
        return "PRE_REPAIR_STAGING"
    if "调梁棚" in targets or "调梁线北" in targets:
        return "DISPATCH_SHED_QUEUE"
    if targets & {"洗罐站", "洗罐线北", "油漆线", "抛丸线", "卸轮线", "机走棚", "机库线", "机走北"}:
        return "FUNCTION_LINE_SERVICE"
    if targets & {"存1线", "存2线", "存3线", "存5线", "存5线北", "存5线南", "存4线"}:
        return "YARD_REBALANCE"
    if line in {"修1库内", "修2库内", "修3库内", "修4库内"} or repair_process:
        return "REPAIR_CONTEXT"
    return None


def audit_truth_case(path: Path) -> TruthCaseAudit:
    case_id = case_id_from_path(path)
    payload = load_truth_case(path)
    cars = payload.get("StartStatus") or []
    movable = [car for car in cars if not is_satisfied(car)]
    families = [contract_family(car) for car in movable]
    covered = sum(1 for item in families if item is not None)
    residual = len(movable) - covered
    vehicle_count = len(cars)
    movable_count = len(movable)
    coverage = covered / movable_count if movable_count else 1.0
    residual_ratio = residual / movable_count if movable_count else 0.0

    target_counter: Counter[str] = Counter()
    line_counter: Counter[str] = Counter()
    unknown_alias = 0
    special = 0
    for car in cars:
        line = current_line(car)
        line_counter[line] += 1
        if not line:
            unknown_alias += 1
        for target in target_lines(car):
            target_counter[target] += 1
        if car.get("IsWeigh") or car.get("IsClosedDoor") or car.get("IsHeavy"):
            special += 1

    failures: list[str] = []
    if coverage < 0.95:
        failures.append("P1_EFFECTIVE_CONTRACT_COVERAGE_BELOW_95")
    if residual_ratio > 0.05:
        failures.append("P1_RESIDUAL_RATIO_ABOVE_5")
    if vehicle_count == 0:
        failures.append("TRUTH_EMPTY_START_STATUS")

    return TruthCaseAudit(
        case_id=case_id,
        source_path=str(path),
        vehicle_count=vehicle_count,
        movable_vehicle_count=movable_count,
        effective_contract_coverage=round(coverage, 6),
        residual_vehicle_ratio=round(residual_ratio, 6),
        depot_target_vehicle_count=sum(
            target_counter[line]
            for line in ("修1库内", "修2库内", "修3库内", "修4库内", "修1库外", "修2库外", "修3库外", "修4库外")
        ),
        depot_initial_vehicle_count=sum(
            line_counter[line] for line in ("修1库内", "修2库内", "修3库内", "修4库内")
        ),
        pre_repair_target_vehicle_count=target_counter["预修线"],
        dispatch_shed_target_vehicle_count=target_counter["调梁棚"] + target_counter["调梁线北"],
        function_line_target_vehicle_count=sum(
            target_counter[line]
            for line in ("洗罐站", "洗罐线北", "油漆线", "抛丸线", "卸轮线", "机走棚", "机库线", "机走北")
        ),
        storage_target_vehicle_count=sum(
            target_counter[line] for line in ("存1线", "存2线", "存3线", "存5线", "存5线北", "存5线南")
        ),
        cun4_target_vehicle_count=target_counter["存4线"],
        special_process_vehicle_count=special,
        line_alias_unknown_count=unknown_alias,
        passed_p1_floor=not failures,
        failures=failures,
    )


REQUIRED_START_STATUS_FIELDS = (
    "Line",
    "Position",
    "RepairProcess",
    "Type",
    "No",
    "Length",
    "IsHeavy",
    "IsWeigh",
    "IsClosedDoor",
    "TargetLines",
)


def audit_p0_online_boundary(path: Path) -> ProcessAudit:
    case_id = case_id_from_path(path)
    payload = load_truth_case(path)
    cars = payload.get("StartStatus") or []
    terminal_lines = payload.get("TerminalLines") or []
    missing_field_count = 0
    unknown_line_count = 0
    duplicate_vehicle_count = 0
    position_collision_count = 0

    vehicle_ids: Counter[str] = Counter()
    positions: Counter[tuple[str, Any]] = Counter()
    for car in cars:
        for field_name in REQUIRED_START_STATUS_FIELDS:
            if field_name not in car or car[field_name] in (None, ""):
                missing_field_count += 1
        line = current_line(car)
        if not line:
            unknown_line_count += 1
        vehicle_ids[str(car.get("No") or "")] += 1
        positions[(line, car.get("Position"))] += 1

    duplicate_vehicle_count = sum(count - 1 for key, count in vehicle_ids.items() if key and count > 1)
    position_collision_count = sum(count - 1 for count in positions.values() if count > 1)
    terminal_set = {normalize_line(item.get("Line")) for item in terminal_lines}
    failures: list[str] = []
    if not cars:
        failures.append("P0_START_STATUS_EMPTY")
    if missing_field_count:
        failures.append("P0_REQUIRED_FIELD_MISSING")
    if unknown_line_count:
        failures.append("P0_UNKNOWN_LINE")
    if duplicate_vehicle_count:
        failures.append("P0_DUPLICATE_VEHICLE_NO")
    if position_collision_count:
        failures.append("P0_POSITION_COLLISION")
    if terminal_set != {"修1库内", "修2库内", "修3库内", "修4库内"}:
        failures.append("P0_TERMINAL_LINES_INVALID")

    return ProcessAudit(
        case_id=case_id,
        process="P0_ONLINE_EVIDENCE_BOUNDARY",
        status="passed" if not failures else "failed",
        checked=True,
        metric_1_name="missing_field_count",
        metric_1_value=str(missing_field_count),
        metric_2_name="duplicate_vehicle_count",
        metric_2_value=str(duplicate_vehicle_count),
        metric_3_name="position_collision_count",
        metric_3_value=str(position_collision_count),
        failures=failures,
    )


def audit_p1_vehicle_classification(truth_audit: TruthCaseAudit) -> ProcessAudit:
    failures = list(truth_audit.failures)
    return ProcessAudit(
        case_id=truth_audit.case_id,
        process="P1_VEHICLE_CLASSIFICATION",
        status="passed" if not failures else "failed",
        checked=True,
        metric_1_name="effective_contract_coverage",
        metric_1_value=str(truth_audit.effective_contract_coverage),
        metric_2_name="residual_vehicle_ratio",
        metric_2_value=str(truth_audit.residual_vehicle_ratio),
        metric_3_name="movable_vehicle_count",
        metric_3_value=str(truth_audit.movable_vehicle_count),
        failures=failures,
    )


def audit_p2_edge_signal_baseline(manual_audit: ManualPhaseAudit) -> ProcessAudit:
    failures: list[str] = []
    if manual_audit.variant == "FULL_CHAIN_REPAIR":
        if manual_audit.signals.get("cun4_release") is None:
            failures.append("P2_CUN4_RELEASE_SIGNAL_MISSING")
        if manual_audit.signals.get("machine_accept_after_release") is None:
            failures.append("P2_MACHINE_ACCEPT_SIGNAL_MISSING")
        if manual_audit.signals.get("depot_digest_after_accept") is None:
            failures.append("P2_DEPOT_DIGEST_SIGNAL_MISSING")
    elif manual_audit.variant == "DEPOT_DIGEST_ONLY":
        if manual_audit.signals.get("depot_digest") is None:
            failures.append("P2_DIGEST_ONLY_SIGNAL_MISSING")
    elif manual_audit.variant == "MIXED_SIGNAL_REPAIR":
        if manual_audit.signals.get("cun4_release") is None:
            failures.append("P2_MIXED_SIGNAL_RELEASE_MISSING")
    elif manual_audit.variant == "DIRECT_REPAIR_ENTRY":
        if manual_audit.signals.get("depot_entry") is None:
            failures.append("P2_DIRECT_ENTRY_DEPOT_SIGNAL_MISSING")

    return ProcessAudit(
        case_id=manual_audit.case_id,
        process="P2_EDGE_SIGNAL_BASELINE",
        status="passed" if not failures else "failed",
        checked=True,
        metric_1_name="variant",
        metric_1_value=manual_audit.variant,
        metric_2_name="phase_path",
        metric_2_value=manual_audit.phase_path,
        metric_3_name="confidence",
        metric_3_value=manual_audit.confidence,
        failures=failures,
    )


def audit_p3_contract_variant_baseline(manual_audit: ManualPhaseAudit) -> ProcessAudit:
    failures = list(manual_audit.failures)
    valid_variants = {
        "FULL_CHAIN_REPAIR",
        "DEPOT_DIGEST_ONLY",
        "MIXED_SIGNAL_REPAIR",
        "DIRECT_REPAIR_ENTRY",
        "GENERAL_ORGANIZATION",
    }
    if manual_audit.variant not in valid_variants:
        failures.append("P3_UNKNOWN_VARIANT")
    if not manual_audit.phase_path:
        failures.append("P3_PHASE_PATH_MISSING")
    return ProcessAudit(
        case_id=manual_audit.case_id,
        process="P3_CONTRACT_VARIANT_BASELINE",
        status="passed" if not failures else "failed",
        checked=True,
        metric_1_name="variant",
        metric_1_value=manual_audit.variant,
        metric_2_name="observed_hook_count",
        metric_2_value=str(manual_audit.observed_hook_count),
        metric_3_name="soft_hook_upper_bound",
        metric_3_value=str(manual_audit.soft_hook_upper_bound),
        failures=failures,
    )


def runtime_blocked_audit(case_id: str, process: str, reason: str) -> ProcessAudit:
    return ProcessAudit(
        case_id=case_id,
        process=process,
        status="blocked_waiting_runtime_trace",
        checked=False,
        metric_1_name="blocker",
        metric_1_value=reason,
        failures=[],
    )


def build_process_audits(
    truth_paths: list[Path],
    truth_audits: list[TruthCaseAudit],
    manual_audits: list[ManualPhaseAudit],
    phase_trace_audits: list[PhaseTraceAudit],
    action_trace_audits: list[ActionTraceAudit],
    candidate_trace_audits: list[CandidateTraceAudit],
    resource_trace_audits: list[ResourceTraceAudit],
    delta_trace_audits: list[DeltaTraceAudit],
    optimization_trace_audits: list[OptimizationTraceAudit],
    state_update_trace_audits: list[StateUpdateTraceAudit],
    candidate_baselines: list[ManualCandidateBaseline],
) -> list[ProcessAudit]:
    process_rows: list[ProcessAudit] = []
    truth_by_case = {audit.case_id: audit for audit in truth_audits}
    manual_by_case: dict[str, ManualPhaseAudit] = {}
    for audit in manual_audits:
        manual_by_case.setdefault(audit.case_id, audit)
    action_by_case: dict[str, list[ActionTraceAudit]] = {}
    for audit in action_trace_audits:
        action_by_case.setdefault(audit.case_id, []).append(audit)
    candidate_by_case: dict[str, list[CandidateTraceAudit]] = {}
    for audit in candidate_trace_audits:
        candidate_by_case.setdefault(audit.case_id, []).append(audit)
    resource_by_case: dict[str, list[ResourceTraceAudit]] = {}
    for audit in resource_trace_audits:
        resource_by_case.setdefault(audit.case_id, []).append(audit)
    delta_by_case: dict[str, list[DeltaTraceAudit]] = {}
    for audit in delta_trace_audits:
        delta_by_case.setdefault(audit.case_id, []).append(audit)
    optimization_by_case: dict[str, list[OptimizationTraceAudit]] = {}
    for audit in optimization_trace_audits:
        optimization_by_case.setdefault(audit.case_id, []).append(audit)
    state_update_by_case: dict[str, list[StateUpdateTraceAudit]] = {}
    for audit in state_update_trace_audits:
        state_update_by_case.setdefault(audit.case_id, []).append(audit)
    candidate_baseline_by_case: dict[str, list[ManualCandidateBaseline]] = {}
    for baseline in candidate_baselines:
        candidate_baseline_by_case.setdefault(baseline.case_id, []).append(baseline)

    for path in truth_paths:
        case_id = case_id_from_path(path)
        process_rows.append(audit_p0_online_boundary(path))
        truth_audit = truth_by_case.get(case_id)
        if truth_audit:
            process_rows.append(audit_p1_vehicle_classification(truth_audit))
        else:
            process_rows.append(runtime_blocked_audit(case_id, "P1_VEHICLE_CLASSIFICATION", "truth_audit_missing"))
        manual_audit = manual_by_case.get(case_id)
        action_audits = action_by_case.get(case_id, [])
        candidate_audits = candidate_by_case.get(case_id, [])
        resource_audits = resource_by_case.get(case_id, [])
        delta_audits = delta_by_case.get(case_id, [])
        optimization_audits = optimization_by_case.get(case_id, [])
        state_update_audits = state_update_by_case.get(case_id, [])
        case_candidate_baselines = candidate_baseline_by_case.get(case_id, [])
        action_failures = [failure for audit in action_audits for failure in audit.failures]
        p2_action_failures = [failure for failure in action_failures if failure.startswith("P2_") or failure == "ACTION_PHASE_CONTRACT_MISSING"]
        p3_action_failures = [failure for failure in action_failures if failure.startswith("P3_")]
        p4_action_failures = [failure for failure in action_failures if failure.startswith("P4_")]

        if action_audits:
            process_rows.append(
                ProcessAudit(
                    case_id=case_id,
                    process="P2_EDGE_SIGNAL_RUNTIME",
                    status="passed" if not p2_action_failures else "failed",
                    checked=True,
                    metric_1_name="action_trace_record_count",
                    metric_1_value=str(len(action_audits)),
                    metric_2_name="p2_failure_count",
                    metric_2_value=str(len(p2_action_failures)),
                    failures=p2_action_failures,
                )
            )
            process_rows.append(
                ProcessAudit(
                    case_id=case_id,
                    process="P3_CONTRACT_RUNTIME",
                    status="passed" if not p3_action_failures else "failed",
                    checked=True,
                    metric_1_name="action_trace_record_count",
                    metric_1_value=str(len(action_audits)),
                    metric_2_name="p3_failure_count",
                    metric_2_value=str(len(p3_action_failures)),
                    failures=p3_action_failures,
                )
            )
        elif manual_audit:
            process_rows.append(audit_p2_edge_signal_baseline(manual_audit))
            process_rows.append(audit_p3_contract_variant_baseline(manual_audit))
        else:
            process_rows.append(runtime_blocked_audit(case_id, "P2_EDGE_SIGNAL_BASELINE", "manual_baseline_missing"))
            process_rows.append(runtime_blocked_audit(case_id, "P3_CONTRACT_VARIANT_BASELINE", "manual_baseline_missing"))

        if action_audits:
            process_rows.append(
                ProcessAudit(
                    case_id=case_id,
                    process="P4_TARGET_CONTRACT_INTENT",
                    status="passed" if not p4_action_failures else "failed",
                    checked=True,
                    metric_1_name="action_trace_record_count",
                    metric_1_value=str(len(action_audits)),
                    metric_2_name="p4_failure_count",
                    metric_2_value=str(len(p4_action_failures)),
                    failures=p4_action_failures,
                )
            )
        else:
            process_rows.append(runtime_blocked_audit(case_id, "P4_TARGET_CONTRACT_INTENT", "solver_action_trace_missing"))

        if candidate_audits:
            generated_keys = {
                (audit.phase, audit.action_family)
                for audit in candidate_audits
                if audit.status == "passed" and audit.candidate_status in {"generated", "kept", "accepted"}
            }
            trace_failures = [
                failure
                for audit in candidate_audits
                for failure in audit.failures
            ]
            missing_critical = [
                f"{baseline.phase}:{baseline.action_family}"
                for baseline in case_candidate_baselines
                if baseline.baseline_level == "critical"
                and (baseline.phase, baseline.action_family) not in generated_keys
            ]
            missing_important = [
                f"{baseline.phase}:{baseline.action_family}"
                for baseline in case_candidate_baselines
                if baseline.baseline_level == "important"
                and (baseline.phase, baseline.action_family) not in generated_keys
            ]
            p5_failures = list(trace_failures)
            p5_failures.extend(f"P5_CRITICAL_CANDIDATE_MISSING:{item}" for item in missing_critical)
            process_rows.append(
                ProcessAudit(
                    case_id=case_id,
                    process="P5_CANDIDATE_GENERATION",
                    status="passed" if not p5_failures else "failed",
                    checked=True,
                    metric_1_name="candidate_trace_record_count",
                    metric_1_value=str(len(candidate_audits)),
                    metric_2_name="missing_critical_count",
                    metric_2_value=str(len(missing_critical)),
                    metric_3_name="missing_important_count",
                    metric_3_value=str(len(missing_important)),
                    failures=p5_failures,
                )
            )
        else:
            process_rows.append(runtime_blocked_audit(case_id, "P5_CANDIDATE_GENERATION", "solver_candidate_trace_missing"))

        if resource_audits:
            candidate_ids = {
                audit.candidate_id
                for audit in candidate_audits
                if audit.status == "passed" and audit.candidate_status in {"generated", "kept", "accepted"}
            }
            resource_candidate_ids = {
                audit.candidate_id
                for audit in resource_audits
                if audit.status == "passed"
            }
            resource_failures = [
                failure
                for audit in resource_audits
                for failure in audit.failures
            ]
            missing_resource = sorted(candidate_ids - resource_candidate_ids)
            resource_status_by_candidate = {
                audit.candidate_id: audit.resource_status
                for audit in resource_audits
                if audit.status == "passed"
            }
            blocked_critical = sorted(
                audit.candidate_id
                for audit in candidate_audits
                if audit.status == "passed"
                and audit.candidate_level == "critical"
                and resource_status_by_candidate.get(audit.candidate_id) in {"blocked", "waiting"}
            )
            p6_failures = list(resource_failures)
            p6_failures.extend(f"P6_RESOURCE_TRACE_MISSING:{item}" for item in missing_resource)
            process_rows.append(
                ProcessAudit(
                    case_id=case_id,
                    process="P6_RESOURCE_ARBITRATION",
                    status="passed" if not p6_failures else "failed",
                    checked=True,
                    metric_1_name="resource_trace_record_count",
                    metric_1_value=str(len(resource_audits)),
                    metric_2_name="missing_resource_count",
                    metric_2_value=str(len(missing_resource)),
                    metric_3_name="blocked_or_waiting_critical_count",
                    metric_3_value=str(len(blocked_critical)),
                    failures=p6_failures,
                )
            )
        else:
            process_rows.append(runtime_blocked_audit(case_id, "P6_RESOURCE_ARBITRATION", "solver_resource_trace_missing"))

        if delta_audits:
            resource_candidate_ids = {
                audit.candidate_id
                for audit in resource_audits
                if audit.status == "passed"
            }
            delta_candidate_ids = {
                audit.candidate_id
                for audit in delta_audits
                if audit.status == "passed"
            }
            delta_failures = [
                failure
                for audit in delta_audits
                for failure in audit.failures
            ]
            missing_delta = sorted(resource_candidate_ids - delta_candidate_ids)
            accepted_hard_violations = [
                audit.candidate_id
                for audit in delta_audits
                if audit.status == "passed"
                and audit.gate_decision == "accept"
                and audit.hard_violation_count
                and audit.hard_violation_count > 0
            ]
            p7_failures = list(delta_failures)
            p7_failures.extend(f"P7_DELTA_TRACE_MISSING:{item}" for item in missing_delta)
            p7_failures.extend(f"P7_HARD_VIOLATION_ACCEPTED:{item}" for item in accepted_hard_violations)
            process_rows.append(
                ProcessAudit(
                    case_id=case_id,
                    process="P7_DELTA_AND_HARD_GATE",
                    status="passed" if not p7_failures else "failed",
                    checked=True,
                    metric_1_name="delta_trace_record_count",
                    metric_1_value=str(len(delta_audits)),
                    metric_2_name="missing_delta_count",
                    metric_2_value=str(len(missing_delta)),
                    metric_3_name="accepted_hard_violation_count",
                    metric_3_value=str(len(accepted_hard_violations)),
                    failures=p7_failures,
                )
            )
        else:
            process_rows.append(runtime_blocked_audit(case_id, "P7_DELTA_AND_HARD_GATE", "solver_delta_trace_missing"))

        if optimization_audits:
            accepted_delta_ids = {
                audit.candidate_id
                for audit in delta_audits
                if audit.status == "passed" and audit.gate_decision == "accept"
            }
            optimized_ids = {
                audit.candidate_id
                for audit in optimization_audits
                if audit.status == "passed"
            }
            optimization_failures = [
                failure
                for audit in optimization_audits
                for failure in audit.failures
            ]
            missing_optimization = sorted(accepted_delta_ids - optimized_ids)
            selected = [
                audit
                for audit in optimization_audits
                if audit.status == "passed" and audit.selected
            ]
            selected_not_rank_one = [
                audit.candidate_id
                for audit in selected
                if audit.rank != 1
            ]
            p8_failures = list(optimization_failures)
            p8_failures.extend(f"P8_OPT_TRACE_MISSING:{item}" for item in missing_optimization)
            p8_failures.extend(f"P8_SELECTED_NOT_RANK_ONE:{item}" for item in selected_not_rank_one)
            if not selected:
                p8_failures.append("P8_NO_SELECTED_CANDIDATE")
            process_rows.append(
                ProcessAudit(
                    case_id=case_id,
                    process="P8_LOCAL_HOOK_OPTIMIZATION",
                    status="passed" if not p8_failures else "failed",
                    checked=True,
                    metric_1_name="optimization_trace_record_count",
                    metric_1_value=str(len(optimization_audits)),
                    metric_2_name="missing_optimization_count",
                    metric_2_value=str(len(missing_optimization)),
                    metric_3_name="selected_candidate_count",
                    metric_3_value=str(len(selected)),
                    failures=p8_failures,
                )
            )
        else:
            process_rows.append(runtime_blocked_audit(case_id, "P8_LOCAL_HOOK_OPTIMIZATION", "solver_optimization_trace_missing"))

        if state_update_audits:
            selected_ids = {
                audit.candidate_id
                for audit in optimization_audits
                if audit.status == "passed" and audit.selected
            }
            updated_ids = {
                audit.candidate_id
                for audit in state_update_audits
                if audit.status == "passed"
            }
            state_failures = [
                failure
                for audit in state_update_audits
                for failure in audit.failures
            ]
            missing_update = sorted(selected_ids - updated_ids)
            hook_increment_total = sum(
                audit.hook_increment or 0
                for audit in state_update_audits
                if audit.status == "passed"
            )
            p9_failures = list(state_failures)
            p9_failures.extend(f"P9_STATE_UPDATE_MISSING:{item}" for item in missing_update)
            process_rows.append(
                ProcessAudit(
                    case_id=case_id,
                    process="P9_STATE_UPDATE_AND_TRACE",
                    status="passed" if not p9_failures else "failed",
                    checked=True,
                    metric_1_name="state_update_trace_record_count",
                    metric_1_value=str(len(state_update_audits)),
                    metric_2_name="missing_update_count",
                    metric_2_value=str(len(missing_update)),
                    metric_3_name="hook_increment_total",
                    metric_3_value=str(hook_increment_total),
                    failures=p9_failures,
                )
            )
        else:
            process_rows.append(runtime_blocked_audit(case_id, "P9_STATE_UPDATE_AND_TRACE", "solver_state_update_trace_missing"))

    if phase_trace_audits:
        failed_cases = {audit.case_id for audit in phase_trace_audits if not audit.passed}
        traced_cases = {audit.case_id for audit in phase_trace_audits}
        for case_id in sorted(traced_cases):
            failures = ["PHASE_TRACE_FAILED"] if case_id in failed_cases else []
            process_rows.append(
                ProcessAudit(
                    case_id=case_id,
                    process="H_PHASE_GATE_RUNTIME_TRACE",
                    status="passed" if not failures else "failed",
                    checked=True,
                    metric_1_name="phase_trace_record_count",
                    metric_1_value=str(sum(1 for audit in phase_trace_audits if audit.case_id == case_id)),
                    metric_2_name="failed_record_count",
                    metric_2_value=str(sum(1 for audit in phase_trace_audits if audit.case_id == case_id and not audit.passed)),
                    failures=failures,
                )
            )

    return process_rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_summary(
    manual_audits: list[ManualPhaseAudit],
    truth_audits: list[TruthCaseAudit],
    skipped_manual_file_count: int,
    skipped_truth_file_count: int,
    phase_contracts: list[PhaseGateContract],
    phase_trace_audits: list[PhaseTraceAudit],
    process_audits: list[ProcessAudit],
    action_trace_audits: list[ActionTraceAudit],
    candidate_baselines: list[ManualCandidateBaseline],
    candidate_trace_audits: list[CandidateTraceAudit],
    resource_trace_audits: list[ResourceTraceAudit],
    delta_trace_audits: list[DeltaTraceAudit],
    optimization_trace_audits: list[OptimizationTraceAudit],
    state_update_trace_audits: list[StateUpdateTraceAudit],
) -> Summary:
    variants = Counter(item.variant for item in manual_audits)
    manual_case_counter = Counter(item.case_id for item in manual_audits)
    manual_case_ids = {item.case_id for item in manual_audits}
    truth_case_ids = {item.case_id for item in truth_audits}
    observed_hooks = [item.observed_hook_count for item in manual_audits if item.observed_hook_count]
    p1_passed = sum(1 for item in truth_audits if item.passed_p1_floor)
    phase_trace_failed = sum(1 for item in phase_trace_audits if not item.passed)
    process_failed = sum(1 for item in process_audits if item.status == "failed")
    process_blocked = sum(1 for item in process_audits if item.status == "blocked_waiting_runtime_trace")
    action_trace_failed = sum(1 for item in action_trace_audits if item.status == "failed")
    candidate_trace_failed = sum(1 for item in candidate_trace_audits if item.status == "failed")
    resource_trace_failed = sum(1 for item in resource_trace_audits if item.status == "failed")
    delta_trace_failed = sum(1 for item in delta_trace_audits if item.status == "failed")
    optimization_trace_failed = sum(1 for item in optimization_trace_audits if item.status == "failed")
    state_update_trace_failed = sum(1 for item in state_update_trace_audits if item.status == "failed")
    return Summary(
        manual_case_count=len(manual_audits),
        truth_case_count=len(truth_audits),
        matched_case_count=len(manual_case_ids & truth_case_ids),
        standard_chain_count=variants["FULL_CHAIN_REPAIR"],
        depot_digest_only_count=variants["DEPOT_DIGEST_ONLY"],
        port_release_without_standard_accept_count=variants["MIXED_SIGNAL_REPAIR"],
        low_signal_or_short_chain_count=variants["DIRECT_REPAIR_ENTRY"],
        median_observed_hooks=median(observed_hooks) if observed_hooks else None,
        p1_passed_case_count=p1_passed,
        p1_failed_case_count=len(truth_audits) - p1_passed,
        skipped_manual_file_count=skipped_manual_file_count,
        skipped_truth_file_count=skipped_truth_file_count,
        duplicate_manual_case_count=sum(count - 1 for count in manual_case_counter.values() if count > 1),
        phase_contract_row_count=len(phase_contracts),
        phase_trace_record_count=len(phase_trace_audits),
        phase_trace_failed_record_count=phase_trace_failed,
        process_audit_row_count=len(process_audits),
        process_failed_row_count=process_failed,
        process_blocked_row_count=process_blocked,
        action_trace_record_count=len(action_trace_audits),
        action_trace_failed_record_count=action_trace_failed,
        candidate_baseline_row_count=len(candidate_baselines),
        candidate_trace_record_count=len(candidate_trace_audits),
        candidate_trace_failed_record_count=candidate_trace_failed,
        resource_trace_record_count=len(resource_trace_audits),
        resource_trace_failed_record_count=resource_trace_failed,
        delta_trace_record_count=len(delta_trace_audits),
        delta_trace_failed_record_count=delta_trace_failed,
        optimization_trace_record_count=len(optimization_trace_audits),
        optimization_trace_failed_record_count=optimization_trace_failed,
        state_update_trace_record_count=len(state_update_trace_audits),
        state_update_trace_failed_record_count=state_update_trace_failed,
    )


def run(args: argparse.Namespace) -> int:
    root = Path(args.root)
    manual_dir = root / args.manual_dir
    truth_dir = root / args.truth_dir
    output_dir = root / args.output_dir
    representative = {case.upper() for case in args.representative_cases}

    manual_paths = sorted(manual_dir.glob("*人工调车作业单/*.xlsx"))
    truth_paths = sorted(truth_dir.glob("validation_*.json"))
    skipped_manual_file_count = sum(1 for path in manual_paths if not try_case_id_from_path(path))
    skipped_truth_file_count = sum(1 for path in truth_paths if not try_case_id_from_path(path))
    manual_paths = [path for path in manual_paths if try_case_id_from_path(path)]
    truth_paths = [path for path in truth_paths if try_case_id_from_path(path)]
    if args.representative_only:
        manual_paths = [path for path in manual_paths if try_case_id_from_path(path) in representative]
        truth_paths = [path for path in truth_paths if try_case_id_from_path(path) in representative]

    manual_audits: list[ManualPhaseAudit] = []
    hooks_by_case: dict[str, list[Hook]] = {}
    for path in manual_paths:
        hooks = parse_manual_hooks(path)
        audit = infer_manual_phase_audit(hooks)
        audit.source_path = str(path)
        manual_audits.append(audit)
        hooks_by_case.setdefault(audit.case_id, hooks)

    truth_audits = [audit_truth_case(path) for path in truth_paths]
    phase_contracts = build_phase_gate_contracts(manual_audits)
    candidate_baselines = build_manual_candidate_baselines(manual_audits, hooks_by_case)
    phase_trace_audits: list[PhaseTraceAudit] = []
    if args.phase_trace:
        phase_trace_audits = audit_phase_trace_records(
            read_phase_trace_records(root / args.phase_trace),
            phase_contracts,
        )
    action_trace_audits: list[ActionTraceAudit] = []
    if args.action_trace:
        action_trace_audits = audit_action_trace_records(
            read_action_trace_records(root / args.action_trace),
            phase_contracts,
        )
    candidate_trace_audits: list[CandidateTraceAudit] = []
    if args.candidate_trace:
        candidate_trace_audits = audit_candidate_trace_records(
            read_candidate_trace_records(root / args.candidate_trace),
            phase_contracts,
            candidate_baselines,
        )
    resource_trace_audits: list[ResourceTraceAudit] = []
    if args.resource_trace:
        resource_trace_audits = audit_resource_trace_records(
            read_resource_trace_records(root / args.resource_trace),
            candidate_trace_audits,
        )
    delta_trace_audits: list[DeltaTraceAudit] = []
    if args.delta_trace:
        delta_trace_audits = audit_delta_trace_records(
            read_delta_trace_records(root / args.delta_trace),
            candidate_trace_audits,
            resource_trace_audits,
        )
    optimization_trace_audits: list[OptimizationTraceAudit] = []
    if args.optimization_trace:
        optimization_trace_audits = audit_optimization_trace_records(
            read_optimization_trace_records(root / args.optimization_trace),
            candidate_trace_audits,
            delta_trace_audits,
        )
    state_update_trace_audits: list[StateUpdateTraceAudit] = []
    if args.state_update_trace:
        state_update_trace_audits = audit_state_update_trace_records(
            read_state_update_trace_records(root / args.state_update_trace),
            optimization_trace_audits,
        )
    process_audits = build_process_audits(
        truth_paths=truth_paths,
        truth_audits=truth_audits,
        manual_audits=manual_audits,
        phase_trace_audits=phase_trace_audits,
        action_trace_audits=action_trace_audits,
        candidate_trace_audits=candidate_trace_audits,
        resource_trace_audits=resource_trace_audits,
        delta_trace_audits=delta_trace_audits,
        optimization_trace_audits=optimization_trace_audits,
        state_update_trace_audits=state_update_trace_audits,
        candidate_baselines=candidate_baselines,
    )
    summary = build_summary(
        manual_audits,
        truth_audits,
        skipped_manual_file_count=skipped_manual_file_count,
        skipped_truth_file_count=skipped_truth_file_count,
        phase_contracts=phase_contracts,
        phase_trace_audits=phase_trace_audits,
        process_audits=process_audits,
        action_trace_audits=action_trace_audits,
        candidate_baselines=candidate_baselines,
        candidate_trace_audits=candidate_trace_audits,
        resource_trace_audits=resource_trace_audits,
        delta_trace_audits=delta_trace_audits,
        optimization_trace_audits=optimization_trace_audits,
        state_update_trace_audits=state_update_trace_audits,
    )

    manual_rows = []
    for audit in manual_audits:
        row = asdict(audit)
        row["failures"] = ";".join(audit.failures)
        row.update({f"signal_{key}": value for key, value in audit.signals.items()})
        row.pop("signals", None)
        manual_rows.append(row)

    truth_rows = []
    for audit in truth_audits:
        row = asdict(audit)
        row["failures"] = ";".join(audit.failures)
        truth_rows.append(row)

    phase_contract_rows = [asdict(contract) for contract in phase_contracts]
    candidate_baseline_rows = [asdict(baseline) for baseline in candidate_baselines]

    phase_trace_rows = []
    for audit in phase_trace_audits:
        row = asdict(audit)
        row["failures"] = ";".join(audit.failures)
        phase_trace_rows.append(row)

    action_trace_rows = []
    for audit in action_trace_audits:
        row = asdict(audit)
        row["failures"] = ";".join(audit.failures)
        action_trace_rows.append(row)

    candidate_trace_rows = []
    for audit in candidate_trace_audits:
        row = asdict(audit)
        row["failures"] = ";".join(audit.failures)
        candidate_trace_rows.append(row)

    resource_trace_rows = []
    for audit in resource_trace_audits:
        row = asdict(audit)
        row["failures"] = ";".join(audit.failures)
        resource_trace_rows.append(row)

    delta_trace_rows = []
    for audit in delta_trace_audits:
        row = asdict(audit)
        row["failures"] = ";".join(audit.failures)
        delta_trace_rows.append(row)

    optimization_trace_rows = []
    for audit in optimization_trace_audits:
        row = asdict(audit)
        row["failures"] = ";".join(audit.failures)
        optimization_trace_rows.append(row)

    state_update_trace_rows = []
    for audit in state_update_trace_audits:
        row = asdict(audit)
        row["failures"] = ";".join(audit.failures)
        state_update_trace_rows.append(row)

    process_rows = []
    for audit in process_audits:
        row = asdict(audit)
        row["failures"] = ";".join(audit.failures)
        process_rows.append(row)

    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "manual_phase_audit.csv", manual_rows)
    write_csv(output_dir / "truth_structure_audit.csv", truth_rows)
    write_csv(output_dir / "phase_gate_contract.csv", phase_contract_rows)
    write_csv(output_dir / "manual_candidate_baseline.csv", candidate_baseline_rows)
    write_csv(output_dir / "process_audit.csv", process_rows)
    if args.phase_trace:
        write_csv(output_dir / "phase_trace_audit.csv", phase_trace_rows)
    if args.action_trace:
        write_csv(output_dir / "action_trace_audit.csv", action_trace_rows)
    if args.candidate_trace:
        write_csv(output_dir / "candidate_trace_audit.csv", candidate_trace_rows)
    if args.resource_trace:
        write_csv(output_dir / "resource_trace_audit.csv", resource_trace_rows)
    if args.delta_trace:
        write_csv(output_dir / "delta_trace_audit.csv", delta_trace_rows)
    if args.optimization_trace:
        write_csv(output_dir / "optimization_trace_audit.csv", optimization_trace_rows)
    if args.state_update_trace:
        write_csv(output_dir / "state_update_trace_audit.csv", state_update_trace_rows)
    (output_dir / "phase_gate_summary.json").write_text(
        json.dumps(asdict(summary), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(json.dumps(asdict(summary), ensure_ascii=False, indent=2))
    print(f"Wrote {output_dir / 'manual_phase_audit.csv'}")
    print(f"Wrote {output_dir / 'truth_structure_audit.csv'}")
    print(f"Wrote {output_dir / 'phase_gate_contract.csv'}")
    print(f"Wrote {output_dir / 'manual_candidate_baseline.csv'}")
    print(f"Wrote {output_dir / 'process_audit.csv'}")
    if args.phase_trace:
        print(f"Wrote {output_dir / 'phase_trace_audit.csv'}")
    if args.action_trace:
        print(f"Wrote {output_dir / 'action_trace_audit.csv'}")
    if args.candidate_trace:
        print(f"Wrote {output_dir / 'candidate_trace_audit.csv'}")
    if args.resource_trace:
        print(f"Wrote {output_dir / 'resource_trace_audit.csv'}")
    if args.delta_trace:
        print(f"Wrote {output_dir / 'delta_trace_audit.csv'}")
    if args.optimization_trace:
        print(f"Wrote {output_dir / 'optimization_trace_audit.csv'}")
    if args.state_update_trace:
        print(f"Wrote {output_dir / 'state_update_trace_audit.csv'}")
    print(f"Wrote {output_dir / 'phase_gate_summary.json'}")
    if args.check:
        errors = check_summary(
            summary,
            manual_audits,
            truth_audits,
            phase_trace_audits,
            process_audits,
            action_trace_audits,
            candidate_trace_audits,
            resource_trace_audits,
            delta_trace_audits,
            optimization_trace_audits,
            state_update_trace_audits,
        )
        if errors:
            for error in errors:
                print(f"CHECK_FAILED: {error}")
            return 1
    return 0


def check_summary(
    summary: Summary,
    manual_audits: list[ManualPhaseAudit],
    truth_audits: list[TruthCaseAudit],
    phase_trace_audits: list[PhaseTraceAudit],
    process_audits: list[ProcessAudit],
    action_trace_audits: list[ActionTraceAudit],
    candidate_trace_audits: list[CandidateTraceAudit],
    resource_trace_audits: list[ResourceTraceAudit],
    delta_trace_audits: list[DeltaTraceAudit],
    optimization_trace_audits: list[OptimizationTraceAudit],
    state_update_trace_audits: list[StateUpdateTraceAudit],
) -> list[str]:
    errors: list[str] = []
    if not manual_audits:
        errors.append("manual audit is empty")
    if not truth_audits:
        errors.append("truth audit is empty")
    if any(item.failures for item in manual_audits):
        errors.append("manual phase failures exist")
    if any(item.failures for item in truth_audits):
        errors.append("truth structure failures exist")
    if summary.manual_case_count >= 90 and summary.standard_chain_count < 90:
        errors.append("standard chain count below expected baseline")
    if summary.manual_case_count < 90 and summary.standard_chain_count == 0:
        errors.append("representative audit has no standard chain cases")
    if summary.p1_failed_case_count != 0:
        errors.append("P1 ownership floor has failing truth cases")
    if any(not item.passed for item in phase_trace_audits):
        errors.append("phase trace failures exist")
    if any(item.status == "failed" for item in action_trace_audits):
        errors.append("action trace failures exist")
    if any(item.status == "failed" for item in candidate_trace_audits):
        errors.append("candidate trace failures exist")
    if any(item.status == "failed" for item in resource_trace_audits):
        errors.append("resource trace failures exist")
    if any(item.status == "failed" for item in delta_trace_audits):
        errors.append("delta trace failures exist")
    if any(item.status == "failed" for item in optimization_trace_audits):
        errors.append("optimization trace failures exist")
    if any(item.status == "failed" for item in state_update_trace_audits):
        errors.append("state update trace failures exist")
    if any(item.status == "failed" for item in process_audits):
        errors.append("process audit failures exist")
    return errors


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate manual H1-H5 phase signals and first-pass truth2 structure coverage.",
    )
    parser.add_argument("--root", default=".", help="Repository root.")
    parser.add_argument("--manual-dir", default="data/人工调车数据")
    parser.add_argument("--truth-dir", default="data/truth2")
    parser.add_argument("--output-dir", default="artifacts/phase_gate_audit")
    parser.add_argument(
        "--phase-trace",
        default="",
        help="Optional solver PhaseGateRecord CSV/JSON path to audit.",
    )
    parser.add_argument(
        "--action-trace",
        default="",
        help="Optional solver action trace CSV/JSON path to audit P2/P3/P4 runtime fields.",
    )
    parser.add_argument(
        "--candidate-trace",
        default="",
        help="Optional solver candidate trace CSV/JSON path to audit P5 candidate generation.",
    )
    parser.add_argument(
        "--resource-trace",
        default="",
        help="Optional solver resource trace CSV/JSON path to audit P6 resource arbitration.",
    )
    parser.add_argument(
        "--delta-trace",
        default="",
        help="Optional solver delta trace CSV/JSON path to audit P7 delta and hard-gate checks.",
    )
    parser.add_argument(
        "--optimization-trace",
        default="",
        help="Optional solver optimization trace CSV/JSON path to audit P8 local hook optimization.",
    )
    parser.add_argument(
        "--state-update-trace",
        default="",
        help="Optional solver state update trace CSV/JSON path to audit P9 state update and rebuild.",
    )
    parser.add_argument(
        "--representative-cases",
        nargs="*",
        default=list(DEFAULT_REPRESENTATIVE_CASES),
        help="Case ids used when --representative-only is set.",
    )
    parser.add_argument("--representative-only", action="store_true")
    parser.add_argument(
        "--check",
        action="store_true",
        help="Fail if current baseline audit invariants are not met.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    raise SystemExit(run(parse_args()))
