#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl

from solver_vnext import physical


EXPECTED_HEADER = ("顺序", "股道", "作业方法", "辆数")
MANUAL_BLOCKS = ((0, "left"), (5, "right"))
SEMANTIC_COUNT_TOKENS = {"回", "停"}
DUI_NOTE_PATTERN = re.compile(r"对\d+")
DAI_NOTE_PATTERN = re.compile(r"代\d+")
DING_NOTE_PATTERN = re.compile(r"顶\d+")
FORK_CONNECT_PATTERN = re.compile(r"叉线接")
FORK_DETACH_PATTERN = re.compile(r"叉线.*摘")
MERGED_NOTE_PATTERNS = (
    DUI_NOTE_PATTERN,
    DAI_NOTE_PATTERN,
    DING_NOTE_PATTERN,
    re.compile(r"接代"),
    FORK_CONNECT_PATTERN,
    FORK_DETACH_PATTERN,
)
CAR_NO_PATTERN = re.compile(r"(?<!\d)\d{7}(?!\d)")
POSITIONING_PATTERN = re.compile(r"(对\d+|顶\d+|连续对位)")
AMBIGUOUS_LINE_ALIASES = {"存5", "注意存5", "存5线", "库", "注意库"}
METHODS = {"+", "-", "顶", "称", "代", "外", "回"}
SPOTTING_POSITIONING_LINES = {"调梁棚", "洗罐站", "油漆线", "抛丸线"}
DEPOT_SHORTHAND_PATTERN = re.compile(r"^(?:注意)?修([1-4])$")
STORAGE5_ALIASES = {"存5", "注意存5", "存5线"}
EXPERT_EXCLUDED_CASES = {
    "0103W": "field_expert_confirmed_holiday_merged_hook_plan_not_reference",
}


@dataclass(frozen=True)
class ManualOperation:
    manual_file: str
    manual_file_id: str
    case_id: str
    shift: str
    hook: int
    block: str
    excel_row: int
    line_raw: str
    line: str
    method_raw: str
    method: str
    count_raw: str
    count: int | None
    note: str


@dataclass(frozen=True)
class ValidationIssue:
    manual_file: str
    manual_file_id: str
    case_id: str
    scope: str
    manual_hook: str
    severity: str
    issue_code: str
    detail: str


@dataclass(frozen=True)
class LineResolution:
    resolved_line: str
    candidates: tuple[str, ...]
    status: str
    reason: str
    confidence: str


@dataclass(frozen=True)
class CountResolution:
    effective_count: int | None
    status: str
    reason: str
    confidence: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Clean and validate manual shunting action labels.")
    parser.add_argument("--root", default=".")
    parser.add_argument("--manual-root", default="data/人工调车数据")
    parser.add_argument("--truth-dir", default="data/truth2")
    parser.add_argument("--output-dir", required=True)
    return parser.parse_args()


def run(args: argparse.Namespace) -> int:
    root = Path(args.root)
    manual_root = root / args.manual_root
    truth_dir = root / args.truth_dir
    output_dir = root / args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    truth_by_case = {
        physical.case_id_from_path(path): path
        for path in sorted(truth_dir.glob("*.json"))
        if path.name != "conversion_summary.json"
    }

    all_operations: list[ManualOperation] = []
    all_issues: list[ValidationIssue] = []
    case_summaries: list[dict[str, Any]] = []

    all_xlsx_paths = sorted(manual_root.glob("*-人工调车作业单/*.xlsx"))
    manual_paths = [path for path in all_xlsx_paths if path.name.startswith("调车作业通知单")]
    excluded_files = [
        {
            "manual_file": str(path.relative_to(root)),
            "reason": "filename_not_manual_notice",
        }
        for path in all_xlsx_paths
        if path not in set(manual_paths)
    ]
    for path in manual_paths:
        operations, issues = parse_manual_file(path, root)
        if operations and operations[0].case_id not in truth_by_case:
            first = operations[0]
            issues.append(
                file_issue(
                    first.manual_file,
                    first.manual_file_id,
                    first.case_id,
                    "warn",
                    "truth_case_missing",
                    "cannot run aggregate state validation against truth2",
                )
            )
        all_operations.extend(operations)
        all_issues.extend(issues)
        case_summaries.append(
            build_case_summary(
                operations=operations,
                issues=issues,
                truth_by_case=truth_by_case,
            )
        )

    labels, label_issues = build_clean_labels(all_operations, truth_by_case)
    all_issues.extend(label_issues)
    case_summaries = merge_case_summary_issues(case_summaries, labels, all_issues)
    phase_segments = build_phase_segments(labels)

    write_csv(output_dir / "manual_raw_operations.csv", [asdict(row) for row in all_operations])
    write_csv(output_dir / "manual_action_labels.csv", labels)
    write_csv(output_dir / "manual_phase_segments.csv", phase_segments)
    write_csv(output_dir / "manual_label_validation.csv", [asdict(row) for row in all_issues])
    write_csv(output_dir / "manual_case_label_summary.csv", case_summaries)
    write_csv(output_dir / "manual_excluded_files.csv", excluded_files)

    summary = build_summary(case_summaries, labels, all_issues)
    summary["excluded_file_count"] = len(excluded_files)
    (output_dir / "manual_label_cleaning_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"wrote: {output_dir}")
    return 0 if summary["error_issue_count"] == 0 else 1


def parse_manual_file(path: Path, root: Path) -> tuple[list[ManualOperation], list[ValidationIssue]]:
    manual_file = str(path.relative_to(root))
    manual_file_id = stable_file_id(path, root)
    issues: list[ValidationIssue] = []
    try:
        case_id = physical.case_id_from_path(path)
    except ValueError as exc:
        issues.append(file_issue(manual_file, manual_file_id, "", "error", "case_id_missing", str(exc)))
        return [], issues

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Sheet1" not in workbook.sheetnames:
        issues.append(file_issue(manual_file, manual_file_id, case_id, "error", "sheet1_missing", "|".join(workbook.sheetnames)))
        return [], issues
    sheet = workbook["Sheet1"]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 4:
        issues.append(file_issue(manual_file, manual_file_id, case_id, "error", "header_row_missing", f"row_count={len(rows)}"))
        return [], issues
    header = rows[3]
    for offset, block in MANUAL_BLOCKS:
        actual = tuple(text(header[offset + index] if len(header) > offset + index else "") for index in range(4))
        if actual != EXPECTED_HEADER:
            issues.append(
                file_issue(
                    manual_file,
                    manual_file_id,
                    case_id,
                    "error",
                    "header_mismatch",
                    f"block={block};actual={actual}",
                )
            )
    if any(issue.issue_code == "header_mismatch" for issue in issues):
        return [], issues

    shift = parse_shift(rows)
    operations: list[ManualOperation] = []
    seen_hooks: Counter[int] = Counter()
    for excel_row, row in enumerate(rows, start=1):
        for offset, block in MANUAL_BLOCKS:
            sequence = row[offset] if len(row) > offset else None
            if not is_positive_int(sequence):
                continue
            hook = int(sequence)
            line_raw = text(row[offset + 1] if len(row) > offset + 1 else "")
            method_raw = text(row[offset + 2] if len(row) > offset + 2 else "")
            count_raw = text(row[offset + 3] if len(row) > offset + 3 else "")
            note = text(row[offset + 4] if len(row) > offset + 4 else "")
            if not (line_raw or method_raw or count_raw or note):
                continue
            seen_hooks[hook] += 1
            count = parse_count(count_raw)
            operations.append(
                ManualOperation(
                    manual_file=manual_file,
                    manual_file_id=manual_file_id,
                    case_id=case_id,
                    shift=shift,
                    hook=hook,
                    block=block,
                    excel_row=excel_row,
                    line_raw=line_raw,
                    line=physical.normalize_line(line_raw),
                    method_raw=method_raw,
                    method=method_raw,
                    count_raw=count_raw,
                    count=count,
                    note=note,
                )
            )

    if not operations:
        issues.append(file_issue(manual_file, manual_file_id, case_id, "error", "no_manual_operations", "no sequence rows parsed"))
        return [], issues

    hooks = sorted(row.hook for row in operations)
    expected = set(range(hooks[0], hooks[-1] + 1))
    missing = sorted(expected - set(hooks))
    for hook in missing:
        issues.append(
            ValidationIssue(
                manual_file=manual_file,
                manual_file_id=manual_file_id,
                case_id=case_id,
                scope="sequence",
                manual_hook=str(hook),
                severity="warn",
                issue_code="sequence_gap_possible_omitted_hook",
                detail=f"expected_range={hooks[0]}..{hooks[-1]}",
            )
        )
    for hook, count in sorted(seen_hooks.items()):
        if count > 1:
            issues.append(hook_issue(manual_file, manual_file_id, case_id, hook, "error", "duplicate_hook_number", f"count={count}"))

    return sorted(operations, key=lambda row: (row.case_id, row.hook, row.block, row.excel_row)), issues


def build_clean_labels(
    operations: list[ManualOperation],
    truth_by_case: dict[str, Path],
) -> tuple[list[dict[str, Any]], list[ValidationIssue]]:
    labels: list[dict[str, Any]] = []
    issues: list[ValidationIssue] = []
    ops_by_file: dict[str, list[ManualOperation]] = {}
    for operation in operations:
        ops_by_file.setdefault(operation.manual_file_id, []).append(operation)

    for manual_file_id in sorted(ops_by_file):
        file_ops = sorted(ops_by_file[manual_file_id], key=lambda row: (row.hook, row.block, row.excel_row))
        case_id = file_ops[0].case_id
        truth_path = truth_by_case.get(case_id)
        aggregate_loads = load_initial_line_counts(truth_path) if truth_path else Counter()
        aggregate_reliable = bool(truth_path)
        records: list[dict[str, Any]] = []
        for operation in file_ops:
            line_resolution = resolve_line(operation, aggregate_loads, aggregate_reliable)
            row_issues = validate_operation(operation, line_resolution, aggregate_loads, aggregate_reliable, truth_path)
            issues.extend(row_issues)
            aggregate_reliable = update_aggregate_state(
                operation,
                line_resolution,
                aggregate_loads,
                aggregate_reliable,
                row_issues,
            )
            records.append(
                {
                    "operation": operation,
                    "line_resolution": line_resolution,
                    "issues": row_issues,
                }
            )
        enrichments, enrichment_issues_by_key = build_rule_enrichments(records)
        issues.extend(
            issue
            for row_issues in enrichment_issues_by_key.values()
            for issue in row_issues
        )

        for record in records:
            operation = record["operation"]
            line_resolution = record["line_resolution"]
            count_resolution = resolve_count(operation)
            row_issues = [
                *record["issues"],
                *enrichment_issues_by_key.get(operation_key(operation), []),
            ]
            phase_label, family, intent, rationale = propose_labels(operation, line_resolution.resolved_line)
            issue_codes = tuple(issue.issue_code for issue in row_issues)
            quality = quality_for(row_issues)
            structural_quality = quality_for(
                [
                    issue
                    for issue in row_issues
                    if issue.issue_code != "car_no_annotation_missing"
                ]
            )
            enrichment = enrichments[operation_key(operation)]
            compound = classify_compound_hook(operation)
            site_semantics = classify_site_semantics(operation, line_resolution.resolved_line)
            reference_scope, reference_reason = manual_reference_scope(operation)
            labels.append(
                {
                    "case_id": operation.case_id,
                    "manual_file_id": operation.manual_file_id,
                    "manual_file": operation.manual_file,
                    "shift": operation.shift,
                    "manual_hook": operation.hook,
                    "source_row": operation.excel_row,
                    "block": operation.block,
                    "line_raw": operation.line_raw,
                    "line": operation.line,
                    "resolved_line": line_resolution.resolved_line,
                    "line_candidates": "|".join(line_resolution.candidates),
                    "line_resolution_status": line_resolution.status,
                    "line_inference_reason": line_resolution.reason,
                    "line_inference_confidence": line_resolution.confidence,
                    "method_raw": operation.method_raw,
                    "method": operation.method,
                    "count_raw": operation.count_raw,
                    "count": operation.count if operation.count is not None else "",
                    "effective_count": count_resolution.effective_count if count_resolution.effective_count is not None else "",
                    "count_resolution_status": count_resolution.status,
                    "count_inference_reason": count_resolution.reason,
                    "count_inference_confidence": count_resolution.confidence,
                    "note": operation.note,
                    "phase_label": phase_label,
                    "family": family,
                    "intent": intent,
                    "owner_nos": "",
                    "blocker_nos": "",
                    "temporary_lines": temporary_lines(operation),
                    "resource_owner": resource_owner(line_resolution.resolved_line),
                    "expected_delta": expected_delta(operation),
                    "manual_rationale": rationale,
                    "label_quality": quality,
                    "structural_label_quality": structural_quality,
                    "annotation_label_quality": "review" if "car_no_annotation_missing" in issue_codes else "clean",
                    "issue_codes": "|".join(issue_codes),
                    "requires_review": int(quality != "clean"),
                    "requires_structural_review": int(structural_quality != "clean"),
                    "manual_reference_scope": reference_scope,
                    "manual_reference_exclusion_reason": reference_reason,
                    **site_semantics,
                    **compound,
                    "vehicle_identity_status": vehicle_identity_status(enrichment),
                    "state_validation_status": "aggregate_checked" if truth_path else "truth_missing",
                    **enrichment,
                }
            )
    return labels, issues


def resolve_line(
    operation: ManualOperation,
    aggregate_loads: Counter[str],
    aggregate_reliable: bool,
) -> LineResolution:
    raw = operation.line_raw
    normalized = operation.line
    if not raw:
        return LineResolution("", tuple(), "unknown", "line_raw_missing", "none")

    depot_match = DEPOT_SHORTHAND_PATTERN.fullmatch(raw)
    if depot_match:
        index = depot_match.group(1)
        inner = f"修{index}库内"
        outer = f"修{index}库外"
        candidates = (inner, outer)
        if "库外" in operation.note:
            return LineResolution(outer, candidates, "inferred", "depot_note_outer", "high")
        reason = "depot_notice_shorthand_inner_default" if raw.startswith("注意") else "depot_shorthand_inner_default"
        return LineResolution(inner, candidates, "inferred", reason, "medium")

    if raw == "留道口洗":
        candidates = ("洗罐站", "洗罐线北", "洗油北")
        return LineResolution("洗罐站", candidates, "inferred", "wash_crossing_shorthand", "medium")

    if raw in STORAGE5_ALIASES:
        candidates = ("存5线北", "存5线南")
        if "南" in operation.note and "北" not in operation.note:
            return LineResolution("存5线南", candidates, "inferred", "note_south_endpoint", "high")
        if "北" in operation.note:
            return LineResolution("存5线北", candidates, "inferred", "note_north_endpoint", "high")
        if aggregate_reliable and operation.method == "+" and operation.count is not None:
            viable = tuple(line for line in candidates if aggregate_loads[line] >= operation.count)
            if len(viable) == 1:
                return LineResolution(viable[0], candidates, "inferred", "unique_aggregate_candidate", "medium")
        return LineResolution("", candidates, "ambiguous", "storage5_endpoint_missing", "none")

    if raw in {"库", "注意库"}:
        return LineResolution("机库线", ("机库线",), "ambiguous", "semantic_yard_shorthand", "low")

    if normalized in physical.TRACK_SPECS:
        full_name = physical.LINE_FULL_NAMES.get(normalized, normalized)
        status = "exact" if raw in {normalized, full_name} else "alias_exact"
        reason = "track_spec_exact" if status == "exact" else "runtime_alias"
        return LineResolution(normalized, (normalized,), status, reason, "high")

    return LineResolution("", (normalized,) if normalized else tuple(), "unknown", "not_in_track_spec", "none")


def resolve_count(operation: ManualOperation) -> CountResolution:
    if operation.count is not None:
        return CountResolution(operation.count, "exact", "numeric_count_cell", "high")
    if operation.count_raw in SEMANTIC_COUNT_TOKENS:
        return CountResolution(None, "semantic", f"semantic_count_cell:{operation.count_raw}", "none")
    if operation.count_raw:
        return CountResolution(None, "unknown", f"non_numeric_count_cell:{operation.count_raw}", "none")
    match = re.fullmatch(r"库外(\d+)(摘)?", operation.note)
    if operation.method == "-" and match:
        suffix = "_with_disconnect" if match.group(2) else ""
        return CountResolution(int(match.group(1)), "inferred", f"kuwai_note_count{suffix}", "high")
    if operation.method in {"+", "-"}:
        return CountResolution(None, "omitted", "blank_count_cell", "none")
    return CountResolution(None, "not_applicable", "method_without_vehicle_count", "none")


def build_rule_enrichments(
    records: list[dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], dict[str, list[ValidationIssue]]]:
    enrichments = {
        operation_key(record["operation"]): empty_rule_enrichment(record["operation"])
        for record in records
    }
    issues_by_key: dict[str, list[ValidationIssue]] = {}
    graph = physical.TrackGraph()

    def add_issue(operation: ManualOperation, code: str, detail: str) -> None:
        issues_by_key.setdefault(operation_key(operation), []).append(
            hook_issue(
                operation.manual_file,
                operation.manual_file_id,
                operation.case_id,
                operation.hook,
                "warn",
                code,
                detail,
            )
        )

    previous_line = ""
    for record in records:
        operation: ManualOperation = record["operation"]
        resolution: LineResolution = record["line_resolution"]
        key = operation_key(operation)
        enrichment = enrichments[key]
        note = operation.note
        car_numbers = tuple(CAR_NO_PATTERN.findall(note))
        enrichment["car_no_values"] = "|".join(car_numbers)
        enrichment["car_no_annotation_present"] = int(bool(car_numbers))
        enrichment["air_hose_connect_present"] = int("接" in note)
        enrichment["air_hose_disconnect_present"] = int("摘" in note)
        enrichment["continuous_positioning_present"] = int(bool(POSITIONING_PATTERN.search(note)))
        enrichment["yard_positioning_present"] = int(bool(POSITIONING_PATTERN.search(note)))
        enrichment["route_switch_present"] = int("叉" in note or any(token in note for token in ("渡", "联")))

        if operation.method == "+" and operation.count is not None:
            enrichment["car_no_annotation_required"] = 1
            enrichment["car_no_annotation_rule"] = "couple_farthest_tail_car"
        elif operation.method == "-" and operation.count is not None and operation.count >= 7:
            enrichment["car_no_annotation_required"] = 1
            enrichment["car_no_annotation_rule"] = "detach_nearest_head_car"
        if enrichment["car_no_annotation_required"] and not enrichment["car_no_annotation_present"]:
            add_issue(operation, "car_no_annotation_missing", enrichment["car_no_annotation_rule"])

        line = resolution.resolved_line
        if operation.method == "-" and line in physical.DEPOT_LINES:
            enrichment["continuous_positioning_required"] = 1
        if operation.method == "+" and line in physical.DEPOT_LINES:
            if operation.shift == "晚上":
                enrichment["continuous_coupling_required"] = 1
            elif operation.shift == "中午":
                enrichment["forklift_assist_expected"] = 1
        if operation.method in {"+", "-"} and line in SPOTTING_POSITIONING_LINES:
            enrichment["yard_positioning_required"] = 1
            if not enrichment["yard_positioning_present"]:
                enrichment["yard_positioning_status"] = "inferred_required"

        if previous_line and line:
            path = graph.route(previous_line, line)
            enrichment["route_passby_path"] = "|".join(path)
            switch_nodes = tuple(node for node in path if node in physical.RUNNING_LINES)
            enrichment["route_switch_nodes"] = "|".join(switch_nodes)
            enrichment["route_switch_required"] = int(bool(switch_nodes))
            enrichment["route_validation_status"] = "map_path_found" if path else "map_path_missing"
            if previous_line != line and not path:
                add_issue(operation, "route_missing_between_resolved_lines", f"{previous_line}->{line}")
        elif line:
            enrichment["route_validation_status"] = "first_resolved_line"
        else:
            enrichment["route_validation_status"] = "line_unresolved"
        if line:
            previous_line = line

    infer_air_hose_rules(records, enrichments, issues_by_key)
    return enrichments, issues_by_key


def empty_rule_enrichment(operation: ManualOperation) -> dict[str, Any]:
    return {
        "car_no_annotation_required": 0,
        "car_no_annotation_present": 0,
        "car_no_annotation_rule": "",
        "car_no_values": "",
        "air_hose_connect_required": 0,
        "air_hose_connect_present": int("接" in operation.note),
        "air_hose_connect_source_hook": "",
        "air_hose_disconnect_required": 0,
        "air_hose_disconnect_present": int("摘" in operation.note),
        "air_hose_state_status": "not_required",
        "consist_count_before": "",
        "consist_count_after": "",
        "continuous_positioning_required": 0,
        "continuous_positioning_present": 0,
        "continuous_coupling_required": 0,
        "forklift_assist_expected": 0,
        "yard_positioning_required": 0,
        "yard_positioning_present": 0,
        "yard_positioning_status": "",
        "route_passby_path": "",
        "route_switch_nodes": "",
        "route_switch_required": 0,
        "route_switch_present": 0,
        "route_validation_status": "",
    }


def infer_air_hose_rules(
    records: list[dict[str, Any]],
    enrichments: dict[str, dict[str, Any]],
    issues_by_key: dict[str, list[ValidationIssue]],
) -> None:
    consist: list[dict[str, Any]] = []
    total = 0
    connected_key = ""
    state_confident = True

    def add_issue(operation: ManualOperation, code: str, detail: str) -> None:
        issues_by_key.setdefault(operation_key(operation), []).append(
            hook_issue(operation.manual_file, operation.manual_file_id, operation.case_id, operation.hook, "warn", code, detail)
        )

    operations_by_key = {
        operation_key(record["operation"]): record["operation"]
        for record in records
    }

    for record in records:
        operation: ManualOperation = record["operation"]
        key = operation_key(operation)
        enrichment = enrichments[key]
        if not state_confident:
            enrichment["air_hose_state_status"] = "uncertain_after_state_gap"
            continue

        if operation.method == "+" and operation.count is not None:
            before = total
            consist.append({"key": key, "hook": operation.hook, "count": operation.count})
            total += operation.count
            enrichment["consist_count_before"] = before
            enrichment["consist_count_after"] = total
            if not connected_key and before < 10 <= total and consist:
                source = consist[0]
                source_key = str(source["key"])
                source_operation = operations_by_key[source_key]
                source_enrichment = enrichments[source_key]
                source_enrichment["air_hose_connect_required"] = 1
                source_enrichment["air_hose_connect_source_hook"] = str(source["hook"])
                source_enrichment["air_hose_state_status"] = "connect_required_from_threshold"
                connected_key = source_key
                if not source_enrichment["air_hose_connect_present"]:
                    add_issue(source_operation, "air_hose_connect_missing", f"threshold_hook={operation.hook};total={total}")
        elif operation.method == "-" and operation.count is not None:
            before = total
            enrichment["consist_count_before"] = before
            if operation.count > total:
                enrichment["air_hose_state_status"] = "uncertain_consist_underflow"
                state_confident = False
                continue
            full_detach = operation.count == total
            remaining = operation.count
            while remaining > 0 and consist:
                segment = consist[-1]
                segment_count = int(segment["count"])
                if segment_count <= remaining:
                    remaining -= segment_count
                    consist.pop()
                else:
                    segment["count"] = segment_count - remaining
                    remaining = 0
            total -= operation.count
            enrichment["consist_count_after"] = total
            if connected_key and full_detach:
                enrichment["air_hose_disconnect_required"] = 1
                enrichment["air_hose_state_status"] = "disconnect_required_for_connected_head"
                if not enrichment["air_hose_disconnect_present"]:
                    add_issue(operation, "air_hose_disconnect_missing", f"connected_source_hook={operations_by_key[connected_key].hook}")
                connected_key = ""
            elif connected_key and operation.count > 0:
                enrichment["air_hose_state_status"] = "partial_detach_head_unknown"
        elif operation.method in {"+", "-"} or operation.method in {"顶", "代", "外", "回", "称"}:
            enrichment["air_hose_state_status"] = "uncertain_state_gap"
            state_confident = False


def validate_operation(
    operation: ManualOperation,
    line_resolution: LineResolution,
    aggregate_loads: Counter[str],
    aggregate_reliable: bool,
    truth_path: Path | None,
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []

    def add(severity: str, code: str, detail: str) -> None:
        issues.append(
            hook_issue(
                operation.manual_file,
                operation.manual_file_id,
                operation.case_id,
                operation.hook,
                severity,
                code,
                detail,
            )
        )

    if not operation.line_raw:
        add("error", "line_missing", "")
    elif line_resolution.status == "unknown":
        add("warn", "line_not_exact_track_spec", f"raw={operation.line_raw};normalized={operation.line}")
    elif line_resolution.status == "ambiguous":
        add(
            "warn",
            "line_resolution_ambiguous",
            f"raw={operation.line_raw};candidates={'|'.join(line_resolution.candidates)};reason={line_resolution.reason}",
        )
        if operation.line_raw in AMBIGUOUS_LINE_ALIASES:
            add("warn", "ambiguous_line_alias", f"raw={operation.line_raw};normalized={operation.line}")
    if not operation.method:
        add("warn", "method_missing", "")
    elif operation.method not in METHODS:
        add("warn", "method_unknown", operation.method)
    if operation.count is None:
        if operation.method in {"+", "-"}:
            add("warn", "count_omitted", f"method={operation.method};raw={operation.count_raw}")
        if operation.count_raw in SEMANTIC_COUNT_TOKENS:
            add("warn", "count_cell_contains_semantic_token", operation.count_raw)
        elif operation.count_raw:
            add("warn", "count_non_numeric", operation.count_raw)
    elif operation.count > physical.PULL_LIMIT_EQUIVALENT:
        add("error", "count_exceeds_pull_limit", f"count={operation.count};limit={physical.PULL_LIMIT_EQUIVALENT}")
    if merged_hook_possible(operation):
        add("warn", "merged_hook_possible", f"method={operation.method};count={operation.count_raw};note={operation.note}")
    resolved_line = line_resolution.resolved_line
    if aggregate_reliable and operation.count is not None and resolved_line in physical.TRACK_SPECS:
        projected = projected_line_count(operation, resolved_line, aggregate_loads)
        if projected < 0:
            add("warn", "aggregate_line_count_negative", f"line={resolved_line};before={aggregate_loads[resolved_line]};count={operation.count}")
    return issues


def update_aggregate_state(
    operation: ManualOperation,
    line_resolution: LineResolution,
    aggregate_loads: Counter[str],
    aggregate_reliable: bool,
    issues: list[ValidationIssue],
) -> bool:
    if not aggregate_reliable:
        return False
    issue_codes = {issue.issue_code for issue in issues}
    resolved_line = line_resolution.resolved_line
    if operation.count is None or resolved_line not in physical.TRACK_SPECS:
        return False
    if line_resolution.status not in {"exact", "alias_exact", "inferred"}:
        return False
    if "aggregate_line_count_negative" in issue_codes:
        return False
    if operation.method == "+":
        aggregate_loads[resolved_line] -= operation.count
    elif operation.method == "-":
        aggregate_loads[resolved_line] += operation.count
    else:
        return False
    return True


def projected_line_count(operation: ManualOperation, resolved_line: str, aggregate_loads: Counter[str]) -> int:
    if operation.method == "+":
        return aggregate_loads[resolved_line] - int(operation.count or 0)
    if operation.method == "-":
        return aggregate_loads[resolved_line] + int(operation.count or 0)
    return aggregate_loads[resolved_line]


def propose_labels(operation: ManualOperation, resolved_line: str) -> tuple[str, str, str, str]:
    text_blob = f"{operation.line_raw}{operation.method}{operation.count_raw}{operation.note}"
    line = resolved_line or operation.line
    if line == "存4线" and ("北头摘" in operation.note or "摘" in operation.note):
        return "H3", "REPAIR_INBOUND", "cun4_release_accept", "存4北释放/摘解信号"
    if line == "存4线":
        return "H2", "CUN4_PORT_STAGING", "cun4_port_forming", "存4/北头口成形或保护"
    if line in physical.DEPOT_TARGET_LINES or operation.line_raw.startswith("修") or "库外" in operation.note:
        return "H4", "REPAIR_INBOUND", "depot_inbound_digest", "修库/库外消化信号"
    if "回" in text_blob or "停" in text_blob:
        return "H5", "TAIL_CLOSEOUT", "tail_closeout", "回库/停留收束信号"
    if line in {"洗罐站", "油漆线", "抛丸线", "卸轮线", "调梁棚", "预修线", "机库线"}:
        return "H1", "FUNCTION_LINE_SERVICE", "front_service_direct", "功能线/前场服务信号"
    if "叉线" in operation.note:
        return "H1", "YARD_REBALANCE", "front_shape_before_remote", "叉线/中间缓冲塑形信号"
    if "北头" in operation.note:
        return "H1", "YARD_REBALANCE", "storage_endpoint_positioning", "北头为端别/现场提示，不单独构成存4阶段"
    return "UNKNOWN", "RESIDUAL", "unknown_manual_intent", "人工动作语义不足"


def temporary_lines(operation: ManualOperation) -> str:
    if "叉线" in operation.note:
        return "预修线"
    return ""


def resource_owner(line: str) -> str:
    if line == "存4线":
        return "CUN4_NORTH_BUFFER"
    if line in physical.DEPOT_TARGET_LINES:
        return "DEPOT_SLOT"
    return ""


def expected_delta(operation: ManualOperation) -> str:
    if operation.count is None:
        return ""
    if operation.method in {"+", "-"}:
        return f"manual_count:{operation.count}"
    return ""


def merged_hook_possible(operation: ManualOperation) -> bool:
    if operation.method in {"顶", "称", "代", "外", "回"}:
        return True
    if operation.count_raw in SEMANTIC_COUNT_TOKENS:
        return True
    return any(pattern.search(operation.note) for pattern in MERGED_NOTE_PATTERNS)


def classify_compound_hook(operation: ManualOperation) -> dict[str, Any]:
    hook_class = compound_hook_class(operation)
    replay_mode = compound_identity_replay_mode(operation, hook_class)
    return {
        "compound_hook_class": hook_class,
        "compound_identity_replay_mode": replay_mode,
        "compound_identity_replay_safe": int(replay_mode.startswith("normal_identity_move")),
        "compound_algorithm_hint": compound_algorithm_hint(operation, hook_class),
    }


def compound_hook_class(operation: ManualOperation) -> str:
    if operation.method in {"顶", "称", "代", "外", "回"}:
        return f"semantic_method_{operation.method}"
    if operation.count_raw in SEMANTIC_COUNT_TOKENS:
        return f"semantic_count_{operation.count_raw}"
    note = operation.note
    if FORK_CONNECT_PATTERN.search(note):
        return "forkline_connect"
    if FORK_DETACH_PATTERN.search(note):
        return "forkline_detach"
    if "接代" in note:
        return "couple_substitute"
    if DUI_NOTE_PATTERN.search(note):
        return "positioning_dui_numeric"
    if DING_NOTE_PATTERN.search(note):
        return "push_positioning_ding_numeric"
    if DAI_NOTE_PATTERN.search(note):
        return "substitute_dai_numeric"
    return "none"


def compound_identity_replay_mode(operation: ManualOperation, hook_class: str) -> str:
    if hook_class == "none":
        return "plain_identity_move"
    if operation.method not in {"+", "-"}:
        return "structural_gap_semantic_method"
    if operation.count is None:
        return "structural_gap_count_omitted"
    if hook_class == "positioning_dui_numeric":
        return "normal_identity_move_with_positioning_note"
    if hook_class == "push_positioning_ding_numeric":
        if operation.line_raw in STORAGE5_ALIASES:
            return "structural_gap_storage5_push_reposition"
        return "normal_identity_move_with_push_note"
    if hook_class in {"forkline_connect", "forkline_detach"}:
        return "normal_identity_move_with_forkline_note"
    return "structural_gap_compound_note"


def compound_algorithm_hint(operation: ManualOperation, hook_class: str) -> str:
    if hook_class == "push_positioning_ding_numeric" and operation.line_raw in STORAGE5_ALIASES:
        match = DING_NOTE_PATTERN.search(operation.note)
        pushed = match.group(0).removeprefix("顶") if match else ""
        moved = str(operation.count or "")
        decomposition = f"存5+{pushed}>存5-{pushed}>存5-{moved}" if pushed and moved else ""
        return (
            f"storage5_push_reposition:pushed_north_count={pushed};manual_count={moved};"
            f"field_hint=decompose_as_storage5_internal_reposition;decomposition_hint={decomposition};"
            "target_south_endpoint_unresolved"
        )
    if hook_class == "forkline_connect":
        return "forkline_means_prepair_line:site_name=存2叉"
    if hook_class == "forkline_detach":
        return "forkline_detach_direction_unresolved:site_name=存2叉"
    return ""


def classify_site_semantics(operation: ManualOperation, resolved_line: str) -> dict[str, Any]:
    if operation.line_raw == "存2" and "叉线" in operation.note:
        return {
            "site_confirmed_alias": "存2叉=预修线",
            "site_confirmed_alias_target": "预修线",
            "site_confirmed_alias_confidence": "field_expert_confirmed",
            "north_head_semantics": "not_structural_for_forkline",
        }
    if resolved_line == "存5线北" and "北头" in operation.note:
        return {
            "site_confirmed_alias": "",
            "site_confirmed_alias_target": "",
            "site_confirmed_alias_confidence": "",
            "north_head_semantics": "endpoint_hint_only",
        }
    return {
        "site_confirmed_alias": "",
        "site_confirmed_alias_target": "",
        "site_confirmed_alias_confidence": "",
        "north_head_semantics": "",
    }


def manual_reference_scope(operation: ManualOperation) -> tuple[str, str]:
    reason = EXPERT_EXCLUDED_CASES.get(operation.case_id, "")
    if reason:
        return "exclude_from_algorithm_learning", reason
    return "reference_candidate", ""


def quality_for(issues: list[ValidationIssue]) -> str:
    if any(issue.severity == "error" for issue in issues):
        return "invalid"
    if issues:
        return "review"
    return "clean"


def build_phase_segments(labels: list[dict[str, Any]]) -> list[dict[str, Any]]:
    segments: list[dict[str, Any]] = []
    by_file: dict[str, list[dict[str, Any]]] = {}
    for row in labels:
        by_file.setdefault(str(row["manual_file_id"]), []).append(row)
    for manual_file_id in sorted(by_file):
        rows = sorted(by_file[manual_file_id], key=lambda row: int(row["manual_hook"]))
        if not rows:
            continue
        start = rows[0]
        current_phase = start["phase_label"]
        segment_start = int(start["manual_hook"])
        segment_issues: Counter[str] = Counter()
        for row in rows:
            phase = row["phase_label"]
            if phase != current_phase:
                previous = rows[max(0, rows.index(row) - 1)]
                segments.append(segment_row(start, current_phase, segment_start, int(previous["manual_hook"]), segment_issues))
                start = row
                current_phase = phase
                segment_start = int(row["manual_hook"])
                segment_issues = Counter()
            for code in str(row["issue_codes"]).split("|"):
                if code:
                    segment_issues[code] += 1
        segments.append(segment_row(start, current_phase, segment_start, int(rows[-1]["manual_hook"]), segment_issues))
    return segments


def segment_row(
    start_row: dict[str, Any],
    phase: str,
    start_hook: int,
    end_hook: int,
    issue_counts: Counter[str],
) -> dict[str, Any]:
    return {
        "case_id": start_row["case_id"],
        "manual_file_id": start_row["manual_file_id"],
        "manual_file": start_row["manual_file"],
        "phase_label": phase,
        "start_hook": start_hook,
        "end_hook": end_hook,
        "hook_count": end_hook - start_hook + 1,
        "issue_counts": counter_text(issue_counts),
    }


def build_case_summary(
    *,
    operations: list[ManualOperation],
    issues: list[ValidationIssue],
    truth_by_case: dict[str, Path],
) -> dict[str, Any]:
    if operations:
        first = operations[0]
        case_id = first.case_id
        manual_file = first.manual_file
        manual_file_id = first.manual_file_id
        hook_count = len(operations)
        min_hook = min(row.hook for row in operations)
        max_hook = max(row.hook for row in operations)
    else:
        first_issue = issues[0] if issues else None
        case_id = first_issue.case_id if first_issue else ""
        manual_file = first_issue.manual_file if first_issue else ""
        manual_file_id = first_issue.manual_file_id if first_issue else ""
        hook_count = 0
        min_hook = 0
        max_hook = 0
    issue_counts = Counter(issue.issue_code for issue in issues)
    severity_counts = Counter(issue.severity for issue in issues)
    return {
        "case_id": case_id,
        "manual_file_id": manual_file_id,
        "manual_file": manual_file,
        "truth2_matched": int(case_id in truth_by_case),
        "hook_count": hook_count,
        "min_hook": min_hook,
        "max_hook": max_hook,
        "issue_count": len(issues),
        "error_count": severity_counts.get("error", 0),
        "warn_count": severity_counts.get("warn", 0),
        "issue_counts": counter_text(issue_counts),
        "clean_hook_count": 0,
        "review_hook_count": 0,
        "invalid_hook_count": 0,
    }


def merge_case_summary_issues(
    summaries: list[dict[str, Any]],
    labels: list[dict[str, Any]],
    issues: list[ValidationIssue],
) -> list[dict[str, Any]]:
    issue_by_file: dict[str, list[ValidationIssue]] = {}
    for issue in issues:
        issue_by_file.setdefault(issue.manual_file_id, []).append(issue)
    labels_by_file: dict[str, list[dict[str, Any]]] = {}
    for label in labels:
        labels_by_file.setdefault(str(label["manual_file_id"]), []).append(label)
    merged: list[dict[str, Any]] = []
    for row in summaries:
        manual_file_id = str(row["manual_file_id"])
        file_issues = issue_by_file.get(manual_file_id, [])
        file_labels = labels_by_file.get(manual_file_id, [])
        issue_counts = Counter(issue.issue_code for issue in file_issues)
        severity_counts = Counter(issue.severity for issue in file_issues)
        quality_counts = Counter(str(label["label_quality"]) for label in file_labels)
        updated = dict(row)
        updated["issue_count"] = len(file_issues)
        updated["error_count"] = severity_counts.get("error", 0)
        updated["warn_count"] = severity_counts.get("warn", 0)
        updated["issue_counts"] = counter_text(issue_counts)
        updated["clean_hook_count"] = quality_counts.get("clean", 0)
        updated["review_hook_count"] = quality_counts.get("review", 0)
        updated["invalid_hook_count"] = quality_counts.get("invalid", 0)
        updated["clean_hook_rate"] = round(quality_counts.get("clean", 0) / len(file_labels), 4) if file_labels else 0
        merged.append(updated)
    return sorted(merged, key=lambda item: (str(item["case_id"]), str(item["manual_file_id"])))


def build_summary(
    case_summaries: list[dict[str, Any]],
    labels: list[dict[str, Any]],
    issues: list[ValidationIssue],
) -> dict[str, Any]:
    label_quality = Counter(str(row["label_quality"]) for row in labels)
    structural_quality = Counter(str(row.get("structural_label_quality") or "") for row in labels)
    annotation_quality = Counter(str(row.get("annotation_label_quality") or "") for row in labels)
    issue_counts = Counter(issue.issue_code for issue in issues)
    severity_counts = Counter(issue.severity for issue in issues)
    matched = sum(int(row["truth2_matched"]) for row in case_summaries)
    line_resolution = Counter(str(row.get("line_resolution_status") or "") for row in labels)
    count_resolution = Counter(str(row.get("count_resolution_status") or "") for row in labels)
    compound_hook_class = Counter(str(row.get("compound_hook_class") or "") for row in labels)
    compound_replay_mode = Counter(str(row.get("compound_identity_replay_mode") or "") for row in labels)
    manual_reference_scope_counts = Counter(str(row.get("manual_reference_scope") or "") for row in labels)
    site_alias_counts = Counter(str(row.get("site_confirmed_alias") or "") for row in labels if row.get("site_confirmed_alias"))
    north_head_semantics = Counter(str(row.get("north_head_semantics") or "") for row in labels if row.get("north_head_semantics"))
    air_hose_state = Counter(str(row.get("air_hose_state_status") or "") for row in labels)

    def sum_field(field: str) -> int:
        return sum(int(row.get(field) or 0) for row in labels)

    return {
        "manual_file_count": len(case_summaries),
        "truth2_matched_file_count": matched,
        "manual_action_count": len(labels),
        "clean_action_count": label_quality.get("clean", 0),
        "review_action_count": label_quality.get("review", 0),
        "invalid_action_count": label_quality.get("invalid", 0),
        "clean_action_rate": round(label_quality.get("clean", 0) / len(labels), 4) if labels else 0,
        "validation_issue_count": len(issues),
        "error_issue_count": severity_counts.get("error", 0),
        "warn_issue_count": severity_counts.get("warn", 0),
        "top_issue_counts": dict(issue_counts.most_common(20)),
        "label_quality_counts": dict(label_quality),
        "structural_label_quality_counts": dict(structural_quality),
        "annotation_label_quality_counts": dict(annotation_quality),
        "structural_clean_action_count": structural_quality.get("clean", 0),
        "structural_review_action_count": structural_quality.get("review", 0),
        "structural_clean_action_rate": round(structural_quality.get("clean", 0) / len(labels), 4) if labels else 0,
        "line_resolution_counts": dict(line_resolution),
        "count_resolution_counts": dict(count_resolution),
        "compound_hook_class_counts": dict(compound_hook_class),
        "compound_identity_replay_mode_counts": dict(compound_replay_mode),
        "manual_reference_scope_counts": dict(manual_reference_scope_counts),
        "site_confirmed_alias_counts": dict(site_alias_counts),
        "north_head_semantics_counts": dict(north_head_semantics),
        "air_hose_state_status_counts": dict(air_hose_state),
        "car_no_annotation_required_count": sum_field("car_no_annotation_required"),
        "car_no_annotation_present_count": sum_field("car_no_annotation_present"),
        "air_hose_connect_required_count": sum_field("air_hose_connect_required"),
        "air_hose_disconnect_required_count": sum_field("air_hose_disconnect_required"),
        "continuous_positioning_required_count": sum_field("continuous_positioning_required"),
        "continuous_coupling_required_count": sum_field("continuous_coupling_required"),
        "forklift_assist_expected_count": sum_field("forklift_assist_expected"),
        "yard_positioning_required_count": sum_field("yard_positioning_required"),
        "route_switch_required_count": sum_field("route_switch_required"),
    }


def load_initial_line_counts(truth_path: Path | None) -> Counter[str]:
    if truth_path is None:
        return Counter()
    _case_id, _payload, cars, _depot_assignment, _loco = physical.read_case(truth_path)
    return Counter(str(car["Line"]) for car in cars)


def stable_file_id(path: Path, root: Path) -> str:
    rel = path.relative_to(root).as_posix()
    return re.sub(r"[^0-9A-Za-z]+", "_", rel).strip("_")


def parse_shift(rows: list[tuple[Any, ...]]) -> str:
    for row in rows[:4]:
        for index, value in enumerate(row):
            if text(value) == "班次":
                return text(row[index + 1] if len(row) > index + 1 else "")
    return ""


def operation_key(operation: ManualOperation) -> str:
    return f"{operation.hook}|{operation.block}|{operation.excel_row}"


def vehicle_identity_status(enrichment: dict[str, Any]) -> str:
    if enrichment.get("car_no_annotation_present"):
        return "manual_car_no_present"
    if enrichment.get("car_no_annotation_required"):
        return "required_car_no_missing"
    return "unresolved_manual_no_car_ids"


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_positive_int(value: Any) -> bool:
    return isinstance(value, (int, float)) and int(value) == value and int(value) > 0


def parse_count(value: str) -> int | None:
    if not value:
        return None
    if re.fullmatch(r"\d+", value):
        return int(value)
    return None


def file_issue(
    manual_file: str,
    manual_file_id: str,
    case_id: str,
    severity: str,
    code: str,
    detail: str,
) -> ValidationIssue:
    return ValidationIssue(
        manual_file=manual_file,
        manual_file_id=manual_file_id,
        case_id=case_id,
        scope="file",
        manual_hook="",
        severity=severity,
        issue_code=code,
        detail=detail,
    )


def hook_issue(
    manual_file: str,
    manual_file_id: str,
    case_id: str,
    hook: int,
    severity: str,
    code: str,
    detail: str,
) -> ValidationIssue:
    return ValidationIssue(
        manual_file=manual_file,
        manual_file_id=manual_file_id,
        case_id=case_id,
        scope="hook",
        manual_hook=str(hook),
        severity=severity,
        issue_code=code,
        detail=detail,
    )


def counter_text(counter: Counter[str]) -> str:
    return "|".join(f"{key}:{count}" for key, count in sorted(counter.items()))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    raise SystemExit(run(parse_args()))
