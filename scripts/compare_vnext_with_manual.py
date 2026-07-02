#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import statistics
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl

from solver_vnext import physical


REPRESENTATIVE_CASES = (
    "0117Z",
    "0310W",
    "0103W",
    "0213W",
    "0306W",
    "0128W",
    "0223W",
    "0308W",
    "0329W",
    "0130Z",
    "0201W",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare vNext runtime artifacts with manual shunting plans.")
    parser.add_argument("--artifact-dir", required=True)
    parser.add_argument("--manual-root", default="data/人工调车数据")
    parser.add_argument("--output-dir", default="")
    return parser.parse_args()


def run(args: argparse.Namespace) -> int:
    artifact_dir = Path(args.artifact_dir)
    output_dir = Path(args.output_dir) if args.output_dir else artifact_dir / "manual_compare"
    output_dir.mkdir(parents=True, exist_ok=True)

    manual_plans = read_manual_plans(Path(args.manual_root))
    case_rows = _read_csv(artifact_dir / "case_summary.csv")
    phase_rows = _read_csv(artifact_dir / "phase_gate_records.csv")
    step_rows = _read_csv(artifact_dir / "step_trace.csv")
    structure_rows = _read_csv(artifact_dir / "structure_node_metrics.csv")

    solver_by_case = {row["case_id"]: row for row in case_rows}
    compare_rows = build_case_compare_rows(manual_plans, solver_by_case, phase_rows, step_rows)
    structure_rows_out = build_structure_effectiveness_rows(
        case_rows=case_rows,
        phase_rows=phase_rows,
        step_rows=step_rows,
        structure_rows=structure_rows,
        compare_rows=compare_rows,
    )
    representative_rows = [
        row for row in compare_rows if row["case_id"] in REPRESENTATIVE_CASES
    ]
    summary = build_summary(compare_rows, case_rows, phase_rows, step_rows)

    physical.write_csv(output_dir / "manual_vs_solver_case_compare.csv", compare_rows)
    physical.write_csv(output_dir / "structure_effectiveness.csv", structure_rows_out)
    physical.write_csv(output_dir / "representative_case_compare.csv", representative_rows)
    physical.write_json(output_dir / "manual_compare_summary.json", summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"wrote: {output_dir}")
    return 0


def read_manual_plans(manual_root: Path) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = {}
    for path in sorted(manual_root.glob("*-人工调车作业单/*.xlsx")):
        case_id = _case_id_from_name(path.name)
        if not case_id:
            continue
        operations = _parse_manual_operations(path)
        if not operations:
            continue
        remote_flags = [op["line"] in physical.REMOTE_INTERACTION_LINES for op in operations]
        plan = {
            "case_id": case_id,
            "manual_file": str(path),
            "manual_hook_count": len(operations),
            "manual_remote_business_transition_count": _transition_count(remote_flags),
            "manual_remote_hook_count": sum(remote_flags),
            "manual_cun4_release_hook": _first_hook(operations, _is_cun4_release),
            "manual_machine_accept_hook": _first_hook(operations, _is_machine_accept),
            "manual_first_depot_digest_hook": _first_hook(operations, _is_depot_digest),
            "manual_depot_return_hook": _first_hook(operations, _is_depot_return),
            "manual_phase_signature": _manual_phase_signature(operations),
            "manual_operations": operations,
        }
        current = result.get(case_id)
        if current is None or _manual_plan_sort_key(plan) < _manual_plan_sort_key(current):
            result[case_id] = plan
    return result


def _parse_manual_operations(path: Path) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    by_sequence: dict[int, dict[str, Any]] = {}
    for row in sheet.iter_rows(values_only=True):
        for offset in (0, 5):
            if len(row) <= offset:
                continue
            sequence = row[offset]
            if not isinstance(sequence, (int, float)) or int(sequence) != sequence:
                continue
            line_raw = _text(row[offset + 1] if len(row) > offset + 1 else "")
            method = _text(row[offset + 2] if len(row) > offset + 2 else "")
            count = _text(row[offset + 3] if len(row) > offset + 3 else "")
            note = _text(row[offset + 4] if len(row) > offset + 4 else "")
            if not (line_raw or method or count or note):
                continue
            item = {
                "hook": int(sequence),
                "line_raw": line_raw,
                "line": physical.normalize_line(line_raw),
                "method": method,
                "count": count,
                "note": note,
            }
            existing = by_sequence.get(item["hook"])
            if existing is None or (not existing["line_raw"] and line_raw):
                by_sequence[item["hook"]] = item
    return [by_sequence[key] for key in sorted(by_sequence)]


def build_case_compare_rows(
    manual_plans: dict[str, dict[str, Any]],
    solver_by_case: dict[str, dict[str, str]],
    phase_rows: list[dict[str, str]],
    step_rows: list[dict[str, str]],
) -> list[dict[str, Any]]:
    phase_by_case: dict[str, list[dict[str, str]]] = defaultdict(list)
    step_by_case: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in phase_rows:
        phase_by_case[row["case_id"]].append(row)
    for row in step_rows:
        step_by_case[row["case_id"]].append(row)

    rows: list[dict[str, Any]] = []
    for case_id in sorted(set(manual_plans) & set(solver_by_case)):
        manual = manual_plans[case_id]
        solver = solver_by_case[case_id]
        phases = phase_by_case.get(case_id, [])
        steps = step_by_case.get(case_id, [])
        target_phase_counts = Counter(_predicate(row).get("target_phase", "") for row in phases)
        to_phase_counts = Counter(row["to_phase"] for row in phases)
        selected_templates = Counter(row["template_name"] for row in steps if row.get("template_name"))
        solver_hook = _int(solver.get("hook_count"))
        manual_hook = int(manual["manual_hook_count"])
        solver_remote_transition = _int(solver.get("remote_business_transition_count"))
        manual_remote_transition = int(manual["manual_remote_business_transition_count"])
        rows.append(
            {
                "case_id": case_id,
                "status": solver.get("status", ""),
                "manual_file": manual["manual_file"],
                "manual_hook_count": manual_hook,
                "solver_hook_count": solver_hook,
                "hook_delta": solver_hook - manual_hook,
                "hook_ratio": _ratio(solver_hook, manual_hook),
                "manual_remote_business_transition_count": manual_remote_transition,
                "solver_remote_business_transition_count": solver_remote_transition,
                "remote_transition_delta": solver_remote_transition - manual_remote_transition,
                "manual_remote_hook_count": manual["manual_remote_hook_count"],
                "final_unsatisfied": _int(solver.get("final_unsatisfied")),
                "blocked_reason": solver.get("blocked_reason", ""),
                "manual_cun4_release_hook": manual["manual_cun4_release_hook"],
                "manual_machine_accept_hook": manual["manual_machine_accept_hook"],
                "manual_first_depot_digest_hook": manual["manual_first_depot_digest_hook"],
                "manual_depot_return_hook": manual["manual_depot_return_hook"],
                "manual_phase_signature": manual["manual_phase_signature"],
                "solver_to_phase_counts": _counter_text(to_phase_counts),
                "solver_target_phase_counts": _counter_text(target_phase_counts),
                "solver_has_h3_phase": int(to_phase_counts.get("H3", 0) > 0),
                "solver_target_h3_count": target_phase_counts.get("H3", 0),
                "solver_depot_outbound_session_count": selected_templates.get("depot_outbound_session", 0),
                "solver_remote_session_digest_count": selected_templates.get("remote_session_directional_digest", 0),
                "solver_top_templates": _counter_text(selected_templates, limit=8),
            }
        )
    return rows


def build_structure_effectiveness_rows(
    *,
    case_rows: list[dict[str, str]],
    phase_rows: list[dict[str, str]],
    step_rows: list[dict[str, str]],
    structure_rows: list[dict[str, str]],
    compare_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    case_count = len(case_rows)
    completed = sum(1 for row in case_rows if row.get("status") == "completed")
    target_phase_counts = Counter(_predicate(row).get("target_phase", "") for row in phase_rows)
    to_phase_counts = Counter(row["to_phase"] for row in phase_rows)
    template_counts = Counter(row["template_name"] for row in step_rows if row.get("template_name"))
    phase_template_counts = Counter((row["phase"], row["template_name"]) for row in step_rows if row.get("template_name"))
    release_ready_count = sum(1 for row in structure_rows if str(row.get("cun4_release_ready")) == "True")
    generated = sum(_int(row.get("generated_candidate_count")) for row in structure_rows)
    accepted = sum(_int(row.get("accepted_candidate_count")) for row in structure_rows)
    phase_veto = sum(_int(row.get("phase_veto_count")) for row in structure_rows)
    resource_reject = sum(_int(row.get("resource_violation_count")) for row in structure_rows)
    contract_reject = sum(_int(row.get("contract_reject_count")) for row in structure_rows)
    matched = len(compare_rows)
    hook_ratios = [_float(row["hook_ratio"]) for row in compare_rows if row["hook_ratio"] != ""]
    remote_deltas = [_int(row["remote_transition_delta"]) for row in compare_rows]
    return [
        _structure_row("FullSolve", completed, case_count, "completed_cases", f"{completed}/{case_count}", completed == case_count),
        _structure_row("ManualHookParity", _p50(hook_ratios), matched, "solver/manual hook ratio p50", _distribution_text(hook_ratios), bool(hook_ratios) and _p50(hook_ratios) <= 1.05),
        _structure_row("ManualRemoteTransitionParity", _p50(remote_deltas), matched, "solver-manual remote transition delta p50", _distribution_text(remote_deltas), bool(remote_deltas) and _p50(remote_deltas) <= 1),
        _structure_row("Cun4ReleaseReadyFact", release_ready_count, len(structure_rows), "runtime records with cun4_release_ready", str(release_ready_count), release_ready_count > 0),
        _structure_row("H3ReleaseAcceptPhase", to_phase_counts.get("H3", 0), len(phase_rows), "actual to_phase=H3", _counter_text(to_phase_counts), to_phase_counts.get("H3", 0) > 0),
        _structure_row("H3TargetRecall", target_phase_counts.get("H3", 0), len(phase_rows), "candidate target_phase=H3", _counter_text(target_phase_counts), target_phase_counts.get("H3", 0) > 0),
        _structure_row("DepotOutboundSessionTemplate", template_counts.get("depot_outbound_session", 0), len(step_rows), "selected depot_outbound_session", _counter_text(template_counts, limit=12), template_counts.get("depot_outbound_session", 0) >= max(1, case_count // 4)),
        _structure_row("RemoteSessionDigestTemplate", template_counts.get("remote_session_directional_digest", 0), len(step_rows), "selected remote_session_directional_digest", _counter_text(template_counts, limit=12), template_counts.get("remote_session_directional_digest", 0) > 0),
        _structure_row("CandidateBoundary", accepted, generated, "accepted/generated candidates", f"accepted={accepted};generated={generated};phase_veto={phase_veto};resource_reject={resource_reject};contract_reject={contract_reject}", generated > 0 and accepted > 0),
        _structure_row("H2TemplateShape", sum(count for (phase, _template), count in phase_template_counts.items() if phase == "H2_CUN4_PORT"), len(step_rows), "selected templates while in H2", _counter_text(Counter({template: count for (phase, template), count in phase_template_counts.items() if phase == "H2_CUN4_PORT"}), limit=8), False),
    ]


def build_summary(
    compare_rows: list[dict[str, Any]],
    case_rows: list[dict[str, str]],
    phase_rows: list[dict[str, str]],
    step_rows: list[dict[str, str]],
) -> dict[str, Any]:
    hook_ratios = [_float(row["hook_ratio"]) for row in compare_rows if row["hook_ratio"] != ""]
    hook_deltas = [_int(row["hook_delta"]) for row in compare_rows]
    remote_deltas = [_int(row["remote_transition_delta"]) for row in compare_rows]
    manual_hooks = [_int(row["manual_hook_count"]) for row in compare_rows]
    solver_hooks = [_int(row["solver_hook_count"]) for row in compare_rows]
    manual_remote = [_int(row["manual_remote_business_transition_count"]) for row in compare_rows]
    solver_remote = [_int(row["solver_remote_business_transition_count"]) for row in compare_rows]
    to_phase_counts = Counter(row["to_phase"] for row in phase_rows)
    target_phase_counts = Counter(_predicate(row).get("target_phase", "") for row in phase_rows)
    template_counts = Counter(row["template_name"] for row in step_rows if row.get("template_name"))
    return {
        "case_count": len(case_rows),
        "completed": sum(1 for row in case_rows if row.get("status") == "completed"),
        "blocked": sum(1 for row in case_rows if row.get("status") == "blocked"),
        "matched_manual_case_count": len(compare_rows),
        "manual_hook_distribution": _distribution(manual_hooks),
        "solver_hook_distribution": _distribution(solver_hooks),
        "hook_ratio_distribution": _distribution(hook_ratios),
        "hook_delta_distribution": _distribution(hook_deltas),
        "manual_remote_transition_distribution": _distribution(manual_remote),
        "solver_remote_transition_distribution": _distribution(solver_remote),
        "remote_transition_delta_distribution": _distribution(remote_deltas),
        "to_phase_counts": dict(to_phase_counts),
        "target_phase_counts": dict(target_phase_counts),
        "selected_template_counts": dict(template_counts.most_common(20)),
    }


def _manual_plan_sort_key(plan: dict[str, Any]) -> tuple[int, str]:
    return (int(plan["manual_hook_count"]), str(plan["manual_file"]))


def _manual_phase_signature(operations: list[dict[str, Any]]) -> str:
    events = []
    for op in operations:
        if _is_cun4_release(op):
            events.append(f"H3_RELEASE@{op['hook']}")
        elif _is_machine_accept(op):
            events.append(f"H3_ACCEPT@{op['hook']}")
        elif _is_depot_digest(op):
            events.append(f"H4_DIGEST@{op['hook']}")
        elif _is_depot_return(op):
            events.append(f"H5_RETURN@{op['hook']}")
    return "|".join(events[:12])


def _is_cun4_release(op: dict[str, Any]) -> bool:
    text = _op_text(op)
    return op["line"] == "存4线" and ("北头摘" in text or ("北头" in text and op["method"] == "-"))


def _is_machine_accept(op: dict[str, Any]) -> bool:
    text = _op_text(op)
    return op["line"] == "机库线" and ("接" in text or op["method"] == "+")


def _is_depot_digest(op: dict[str, Any]) -> bool:
    return op["line"] in physical.DEPOT_TARGET_LINES


def _is_depot_return(op: dict[str, Any]) -> bool:
    text = _op_text(op)
    return op["line"] == "机库线" and ("回" in text or "停" in text)


def _first_hook(operations: list[dict[str, Any]], predicate: Any) -> int | str:
    for op in operations:
        if predicate(op):
            return op["hook"]
    return ""


def _op_text(op: dict[str, Any]) -> str:
    return f"{op.get('line_raw', '')} {op.get('method', '')} {op.get('count', '')} {op.get('note', '')}"


def _case_id_from_name(name: str) -> str:
    match = re.search(r"(\d{4}[ZWzw])", name)
    return match.group(1).upper() if match else ""


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists() or path.stat().st_size == 0:
        return []
    with path.open(encoding="utf-8", newline="") as file:
        return list(csv.DictReader(file))


def _predicate(row: dict[str, str]) -> dict[str, str]:
    values = {}
    for part in (row.get("predicate_values") or "").split(";"):
        if "=" in part:
            key, value = part.split("=", 1)
            values[key] = value
    return values


def _transition_count(flags: list[bool]) -> int:
    return sum(1 for left, right in zip(flags, flags[1:]) if left != right)


def _counter_text(counter: Counter[Any], limit: int | None = None) -> str:
    items = counter.most_common(limit)
    return "|".join(f"{key}:{value}" for key, value in items if key != "")


def _distribution(values: list[float | int]) -> dict[str, Any]:
    if not values:
        return {"count": 0}
    return {
        "count": len(values),
        "mean": round(statistics.mean(values), 4),
        "p50": _p50(values),
        "p75": _percentile(values, 0.75),
        "p90": _percentile(values, 0.9),
        "max": max(values),
        "min": min(values),
    }


def _distribution_text(values: list[float | int]) -> str:
    dist = _distribution(values)
    return ";".join(f"{key}={value}" for key, value in dist.items())


def _p50(values: list[float | int]) -> float | int:
    return _percentile(values, 0.5)


def _percentile(values: list[float | int], percentile: float) -> float | int:
    ordered = sorted(values)
    if not ordered:
        return 0
    index = round((len(ordered) - 1) * percentile)
    value = ordered[index]
    return round(value, 4) if isinstance(value, float) else value


def _int(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _float(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _ratio(numerator: int, denominator: int) -> str:
    if denominator == 0:
        return ""
    return f"{numerator / denominator:.4f}"


def _structure_row(name: str, observed: Any, total: int, metric: str, evidence: str, passed: bool) -> dict[str, Any]:
    return {
        "structure": name,
        "status": "pass" if passed else "fail",
        "observed": observed,
        "total": total,
        "metric": metric,
        "evidence": evidence,
    }


if __name__ == "__main__":
    raise SystemExit(run(parse_args()))
