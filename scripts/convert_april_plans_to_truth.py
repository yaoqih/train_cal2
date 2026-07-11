#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT / "data" / "4月-取送车计划-新"
DEFAULT_OUTPUT = ROOT / "data" / "truth3"
DEFAULT_TRUTH_REFERENCE = ROOT / "data" / "truth2"

CASE_ID_RE = re.compile(r"(\d{8}[WZ])", re.IGNORECASE)
REPAIR_LINE_RE = re.compile(r"^修([1-4])库内$")

DEPOT_LINES = [f"修{index}库内" for index in range(1, 5)]
DEPOT_OUTSIDE_LINES = [f"修{index}库外" for index in range(1, 5)]

LINE_ALIASES = {
    "预修": "预修线",
    "预修线": "预修线",
    "机库": "机库线",
    "机库线": "机库线",
    "机北3": "机走北",
    "机走北": "机走北",
    "机棚": "机走棚",
    "机走棚": "机走棚",
    "调北": "调梁线北",
    "调梁线北": "调梁线北",
    "调棚": "调梁棚",
    "调梁棚": "调梁棚",
    "修1": "修1库内",
    "修2": "修2库内",
    "修3": "修3库内",
    "修4": "修4库内",
    "修1库内": "修1库内",
    "修2库内": "修2库内",
    "修3库内": "修3库内",
    "修4库内": "修4库内",
    "修1库外": "修1库外",
    "修2库外": "修2库外",
    "修3库外": "修3库外",
    "修4库外": "修4库外",
    "轮": "卸轮线",
    "卸轮线": "卸轮线",
    "油": "油漆线",
    "油漆线": "油漆线",
    "存1": "存1线",
    "存2": "存2线",
    "存3": "存3线",
    "存4": "存4线",
    "存4北": "存4线",
    "存1线": "存1线",
    "存2线": "存2线",
    "存3线": "存3线",
    "存4线": "存4线",
    "存5北": "存5线北",
    "存5南": "存5线南",
    "存5线北": "存5线北",
    "存5线南": "存5线南",
    "抛": "抛丸线",
    "抛丸线": "抛丸线",
    "洗北": "洗罐线北",
    "洗罐线北": "洗罐线北",
    "洗南": "洗罐站",
    "洗罐站": "洗罐站",
}

# The workbook layout has a fixed number of terminal rows for each line.  This
# is the authoritative conversion-time check for whether direction labels were
# collapsed; physical length validation remains the solver's responsibility.
LINE_SLOT_CAPACITIES = {
    "预修线": 14,
    "机库线": 5,
    "机走北": 7,
    "机走棚": 7,
    "调梁线北": 10,
    "调梁棚": 11,
    "修1库外": 4,
    "修2库外": 4,
    "修3库外": 4,
    "修4库外": 4,
    "修1库内": 7,
    "修2库内": 7,
    "修3库内": 7,
    "修4库内": 7,
    "卸轮线": 4,
    "油漆线": 9,
    "存1线": 9,
    "存2线": 14,
    "存3线": 21,
    "存4线": 25,
    "存5线北": 20,
    "存5线南": 12,
    "抛丸线": 3,
    "洗罐线北": 8,
    "洗罐站": 7,
}

# In the legacy April sheets, an overloaded short-form `洗南` target denotes
# the wash area allocation.  Cars already in wash south and spotting cars stay
# fixed; later contiguous inbound groups overflow to wash north as a unit.
LEGACY_PAIRED_TARGET_OVERFLOW = {
    "洗南": ("洗罐站", "洗罐线北"),
}

REPAIR_PROCESS_ALIASES = {
    "段": "段修",
    "段修": "段修",
    "厂": "厂修",
    "厂修": "厂修",
    "临": "临修",
    "临修": "临修",
    "其他": "其他",
}

# The April workbook numbers each named line locally.  These are the business
# spotting positions from 福州调车业务文档.md, not the older cross-segment
# position offsets found in truth2.
SPOTTING_POSITIONS = {
    "调梁棚": [1, 2, 3, 4],
    "洗罐站": [1, 2, 3],
    "抛丸线": [1, 2],
    "油漆线": [1, 2],
}

# These two types are new in April and therefore cannot be learned from truth2.
APRIL_TYPE_LENGTHS = {
    # KZ70 maintenance specification: vehicle length 12,074 mm.
    "KZ70": 12.1,
    # U60WK has equivalent length 1.1; truth2 records that as 1.1 * 11 m.
    "U60WK": 12.1,
}

TERMINAL_LINES = [
    {"Line": line, "IsInspectionMode": False}
    for line in DEPOT_LINES
]
LOCO_NODE = {"Line": "机库线", "End": "North"}


@dataclass(frozen=True)
class SourceCar:
    source_file: Path
    row: int
    block: str
    line: Any
    position: Any
    car_type: Any
    number: Any
    repair_process: Any
    remark: Any
    target_line: Any
    spotting: Any
    attribute: Any


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert April shunting-plan workbooks to truth2-style API request JSON files."
    )
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--truth-reference", type=Path, default=DEFAULT_TRUTH_REFERENCE)
    return parser.parse_args()


def case_id(path: Path) -> str:
    match = CASE_ID_RE.search(path.name)
    if not match:
        raise ValueError(f"cannot infer case id from {path}")
    return match.group(1).upper()


def normalize_line(value: Any) -> str:
    raw = text(value)
    if raw not in LINE_ALIASES:
        raise ValueError(f"unknown line name: {raw!r}")
    return LINE_ALIASES[raw]


def normalize_repair_process(value: Any) -> str:
    raw = text(value)
    if raw not in REPAIR_PROCESS_ALIASES:
        raise ValueError(f"unknown repair process: {raw!r}")
    return REPAIR_PROCESS_ALIASES[raw]


def positive_int(value: Any, *, field: str, source: SourceCar) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{source.source_file.name}:{source.row} invalid {field}: {value!r}")
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"{source.source_file.name}:{source.row} invalid {field}: {value!r}"
        ) from exc
    if number <= 0:
        raise ValueError(f"{source.source_file.name}:{source.row} invalid {field}: {value!r}")
    return number


def normalize_current_position(line: str, value: Any, source: SourceCar) -> int:
    position = positive_int(value, field="Position", source=source)
    match = REPAIR_LINE_RE.match(line)
    if not match:
        return position
    line_number = int(match.group(1))
    if position >= 100:
        encoded_line = position // 100
        local_position = position % 100
        if encoded_line != line_number or not 1 <= local_position <= 7:
            raise ValueError(
                f"{source.source_file.name}:{source.row} current depot position "
                f"{position} conflicts with {line}"
            )
        return local_position
    if not 1 <= position <= 7:
        raise ValueError(
            f"{source.source_file.name}:{source.row} invalid depot position: {position}"
        )
    return position


def learn_type_lengths(truth_reference: Path) -> dict[str, float]:
    learned: dict[str, float] = {}
    for path in sorted(truth_reference.glob("validation_*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        for car in payload.get("StartStatus") or []:
            car_type = text(car.get("Type"))
            length = float(car.get("Length"))
            previous = learned.setdefault(car_type, length)
            if previous != length:
                raise ValueError(
                    f"truth reference has inconsistent lengths for {car_type}: "
                    f"{previous} and {length}"
                )
    if not learned:
        raise ValueError(f"no type lengths found in {truth_reference}")
    overlap = set(learned) & set(APRIL_TYPE_LENGTHS)
    if overlap:
        raise ValueError(f"April-only type overrides unexpectedly exist in truth2: {sorted(overlap)}")
    learned.update(APRIL_TYPE_LENGTHS)
    return learned


def parse_source_cars(path: Path) -> list[SourceCar]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "起点" not in workbook.sheetnames:
        raise ValueError(f"{path} does not contain sheet '起点'")
    worksheet = workbook["起点"]
    expected_headers = [
        "股道",
        "序号",
        "车型",
        "车号",
        "修程",
        "备注",
        "扣车日期",
        "目标股道",
        "是否对位",
        "车辆属性",
    ]
    for offset in (0, 10):
        headers = [text(worksheet.cell(3, offset + index + 1).value) for index in range(10)]
        if headers != expected_headers:
            raise ValueError(f"{path.name} has unexpected headers in columns {offset + 1}-{offset + 10}")

    cars: list[SourceCar] = []
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, values_only=True),
        start=4,
    ):
        for offset, block in ((0, "A:J"), (10, "K:T")):
            values = row[offset : offset + 10]
            if not text(values[3]):
                continue
            cars.append(
                SourceCar(
                    source_file=path,
                    row=row_index,
                    block=block,
                    line=values[0],
                    position=values[1],
                    car_type=values[2],
                    number=values[3],
                    repair_process=values[4],
                    remark=values[5],
                    target_line=values[7],
                    spotting=values[8],
                    attribute=values[9],
                )
            )
    return cars


def normalize_number(value: Any, source: SourceCar) -> str:
    number = text(value)
    if number.endswith(".0"):
        number = number[:-2]
    if not re.fullmatch(r"\d{7}", number):
        raise ValueError(f"{source.source_file.name}:{source.row} invalid car number: {value!r}")
    return number


def target_for_car(
    source: SourceCar,
    current_line: str,
    corrections: list[str],
) -> tuple[list[str], list[int] | None]:
    raw_target = text(source.target_line)
    raw_spotting = text(source.spotting)

    if raw_target == "大库":
        if raw_spotting not in {"", "是"}:
            raise ValueError(
                f"{source.source_file.name}:{source.row} invalid 大库 spotting value: {raw_spotting!r}"
            )
        return list(DEPOT_LINES), [1, 2, 3, 4, 5]
    if raw_target == "大库外":
        if raw_spotting:
            raise ValueError(
                f"{source.source_file.name}:{source.row} invalid 大库外 spotting value: {raw_spotting!r}"
            )
        return list(DEPOT_OUTSIDE_LINES), None

    if raw_spotting and raw_spotting != "是":
        encoded_position = positive_int(raw_spotting, field="ForceTargetPosition", source=source)
        if encoded_position < 100:
            raise ValueError(
                f"{source.source_file.name}:{source.row} unsupported numeric target position: "
                f"{encoded_position}"
            )
        encoded_line = encoded_position // 100
        local_position = encoded_position % 100
        if encoded_line not in {1, 2, 3, 4} or not 1 <= local_position <= 7:
            raise ValueError(
                f"{source.source_file.name}:{source.row} invalid encoded depot target: "
                f"{encoded_position}"
            )
        target_line = f"修{encoded_line}库内"
        normalized_raw_target = normalize_line(raw_target) if raw_target else ""
        if normalized_raw_target != target_line:
            corrections.append(
                f"{source.source_file.name}:{source.row}:{source.block} "
                f"target {raw_target or '<blank>'!r} corrected to {target_line!r} "
                f"from position {encoded_position}"
            )
        return [target_line], [local_position]

    target_line = normalize_line(raw_target) if raw_target else current_line
    if raw_spotting == "是":
        positions = SPOTTING_POSITIONS.get(target_line)
        if positions is None:
            raise ValueError(
                f"{source.source_file.name}:{source.row} no spotting-position mapping for {target_line}"
            )
        return [target_line], list(positions)
    return [target_line], None


def convert_car(
    source: SourceCar,
    type_lengths: dict[str, float],
    corrections: list[str],
) -> dict[str, Any]:
    line = normalize_line(source.line)
    position = normalize_current_position(line, source.position, source)
    car_type = text(source.car_type)
    if car_type not in type_lengths:
        raise ValueError(
            f"{source.source_file.name}:{source.row} unknown vehicle type length: {car_type!r}"
        )
    number = normalize_number(source.number, source)
    repair_process = normalize_repair_process(source.repair_process)
    remark = text(source.remark)
    attribute = text(source.attribute)
    target_lines, force_positions = target_for_car(source, line, corrections)

    combined_attributes = "|".join(value for value in (remark, attribute) if value)
    car: dict[str, Any] = {
        "Line": line,
        "Position": position,
        "RepairProcess": repair_process,
        "Type": car_type,
        "No": number,
        "Length": type_lengths[car_type],
        "IsHeavy": attribute in {"重", "重车"} or remark == "重" or "重车" in combined_attributes,
        "IsWeigh": "称重" in combined_attributes,
        "IsClosedDoor": "关门" in combined_attributes,
        "TargetLines": target_lines,
    }
    if force_positions:
        car["ForceTargetPosition"] = force_positions
    return car


def split_legacy_paired_targets(
    source_cars: list[SourceCar],
    cars: list[dict[str, Any]],
    corrections: list[str],
) -> None:
    if len(source_cars) != len(cars):
        raise ValueError("source/car count mismatch during paired-target allocation")

    for raw_target, (primary_line, overflow_line) in LEGACY_PAIRED_TARGET_OVERFLOW.items():
        primary_indexes = [
            index
            for index, car in enumerate(cars)
            if car.get("TargetLines") == [primary_line]
        ]
        primary_capacity = LINE_SLOT_CAPACITIES[primary_line]
        if len(primary_indexes) <= primary_capacity:
            continue

        movable_indexes: list[int] = []
        mandatory_indexes: set[int] = set()
        for index in primary_indexes:
            source = source_cars[index]
            car = cars[index]
            current_line = normalize_line(source.line)
            is_legacy_ambiguous = text(source.target_line) == raw_target
            if (
                not is_legacy_ambiguous
                or current_line == primary_line
                or bool(car.get("ForceTargetPosition"))
            ):
                mandatory_indexes.add(index)
            else:
                movable_indexes.append(index)

        if len(mandatory_indexes) > primary_capacity:
            source = source_cars[min(mandatory_indexes)]
            raise ValueError(
                f"{source.source_file.name}: mandatory {primary_line} targets exceed "
                f"slot capacity {primary_capacity}"
            )

        groups: list[list[int]] = []
        for index in movable_indexes:
            source = source_cars[index]
            if groups:
                previous = source_cars[groups[-1][-1]]
                same_run = (
                    normalize_line(previous.line) == normalize_line(source.line)
                    and previous.block == source.block
                    and source.row == previous.row + 1
                )
                if same_run:
                    groups[-1].append(index)
                    continue
            groups.append([index])

        primary_count = len(mandatory_indexes)
        for group in groups:
            if primary_count + len(group) <= primary_capacity:
                primary_count += len(group)
                continue
            for index in group:
                source = source_cars[index]
                cars[index]["TargetLines"] = [overflow_line]
                corrections.append(
                    f"{source.source_file.name}:{source.row}:{source.block} "
                    f"legacy target {raw_target!r} allocated to {overflow_line!r} "
                    f"because {primary_line!r} has {primary_capacity} terminal slots"
                )

        remaining_primary = sum(
            car.get("TargetLines") == [primary_line]
            for car in cars
        )
        if remaining_primary > primary_capacity:
            source = source_cars[primary_indexes[0]]
            raise ValueError(
                f"{source.source_file.name}: unable to split {primary_line} targets "
                f"within {primary_capacity} terminal slots"
            )


def validate_line_slot_capacities(cars: list[dict[str, Any]], path: Path) -> None:
    current_counts = Counter(car["Line"] for car in cars)
    target_counts = Counter(
        car["TargetLines"][0]
        for car in cars
        if len(car.get("TargetLines") or []) == 1
    )
    for line, capacity in LINE_SLOT_CAPACITIES.items():
        if current_counts[line] > capacity:
            raise ValueError(
                f"{path.name}: current line {line} has {current_counts[line]} cars "
                f"but only {capacity} workbook slots"
            )
        if target_counts[line] > capacity:
            raise ValueError(
                f"{path.name}: target line {line} has {target_counts[line]} cars "
                f"but only {capacity} workbook slots"
            )


def validate_payload(payload: dict[str, Any], path: Path) -> None:
    if list(payload) != ["StartStatus", "TerminalLines", "locoNode"]:
        raise ValueError(f"{path.name}: invalid root field order")
    cars = payload["StartStatus"]
    if not cars:
        raise ValueError(f"{path.name}: StartStatus is empty")
    numbers = [car["No"] for car in cars]
    if numbers != sorted(numbers, key=int):
        raise ValueError(f"{path.name}: StartStatus is not sorted by car number")
    if len(numbers) != len(set(numbers)):
        raise ValueError(f"{path.name}: duplicate car numbers")
    locations = [(car["Line"], car["Position"]) for car in cars]
    if len(locations) != len(set(locations)):
        duplicates = [item for item, count in Counter(locations).items() if count > 1]
        raise ValueError(f"{path.name}: duplicate line positions: {duplicates}")
    for car in cars:
        expected_keys = [
            "Line",
            "Position",
            "RepairProcess",
            "Type",
            "No",
            "Length",
            "IsHeavy",
            "IsWeigh",
            "IsClosedDoor",
            "TargetLines",
        ]
        if "ForceTargetPosition" in car:
            expected_keys.append("ForceTargetPosition")
        if list(car) != expected_keys:
            raise ValueError(f"{path.name}:{car['No']} invalid field order")
        if not car["TargetLines"]:
            raise ValueError(f"{path.name}:{car['No']} has no target line")
        if any(not isinstance(value, int) or value <= 0 for value in car.get("ForceTargetPosition", [])):
            raise ValueError(f"{path.name}:{car['No']} has invalid force positions")
    validate_line_slot_capacities(cars, path)
    if payload["TerminalLines"] != TERMINAL_LINES:
        raise ValueError(f"{path.name}: invalid TerminalLines")
    if payload["locoNode"] != LOCO_NODE:
        raise ValueError(f"{path.name}: invalid locoNode")


def convert_file(
    source_path: Path,
    output_dir: Path,
    type_lengths: dict[str, float],
    corrections: list[str],
) -> tuple[Path, int]:
    source_cars = parse_source_cars(source_path)
    cars = [convert_car(source, type_lengths, corrections) for source in source_cars]
    split_legacy_paired_targets(source_cars, cars, corrections)
    cars.sort(key=lambda car: int(car["No"]))
    payload = {
        "StartStatus": cars,
        "TerminalLines": TERMINAL_LINES,
        "locoNode": LOCO_NODE,
    }
    output_path = output_dir / f"validation_取送车计划_{case_id(source_path)}.json"
    validate_payload(payload, output_path)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return output_path, len(cars)


def main() -> int:
    args = parse_args()
    source_paths = sorted(args.source.glob("*.xlsx"), key=case_id)
    if not source_paths:
        raise ValueError(f"no .xlsx files found in {args.source}")
    case_ids = [case_id(path) for path in source_paths]
    if len(case_ids) != len(set(case_ids)):
        raise ValueError(f"duplicate case ids: {case_ids}")

    type_lengths = learn_type_lengths(args.truth_reference)
    args.output.mkdir(parents=True, exist_ok=True)
    corrections: list[str] = []
    generated: list[tuple[Path, int]] = []
    for source_path in source_paths:
        generated.append(convert_file(source_path, args.output, type_lengths, corrections))

    expected_names = {path.name for path, _ in generated}
    unexpected_outputs = sorted(
        path.name
        for path in args.output.glob("*.json")
        if path.name not in expected_names
    )
    if unexpected_outputs:
        raise ValueError(
            f"output directory contains unrelated JSON files: {unexpected_outputs}"
        )

    print(f"generated_files={len(generated)}")
    print(f"generated_cars={sum(count for _, count in generated)}")
    print(f"output_dir={args.output}")
    print(f"target_corrections={len(corrections)}")
    for correction in corrections:
        print(f"CORRECTION {correction}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
